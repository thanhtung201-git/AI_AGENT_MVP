"""
excel_reader.py — Đọc file Excel (.xlsx / .xls)

Hai chế độ:
  read_excel()            — text thuần (backward compat)
  read_excel_structured() — giữ nguyên tọa độ row/col để LLM hiểu layout bất kỳ

Cài đặt: pip install pandas openpyxl
"""

import pandas as pd
import logging
from typing import Optional, List

logger = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 150  # giới hạn gửi LLM tránh tràn token


def read_excel_structured(file_path: str, sheet_filter: Optional[List[str]] = None) -> dict:
    """
    Đọc Excel giữ nguyên tọa độ row/col — để LLM hiểu layout bảng bất kỳ.

    Thay vì flatten thành text thô, mỗi cell được ghi kèm vị trí:
      Row  3: A=Style No | B=Color | C=Qty | D=Total
      Row  4: A=HZSH6C331 | B=Navy | C=100 | D=100
    LLM đọc được cấu trúc 2D → tự hiểu header ở đâu, data ở đâu, dù format nào.

    Args:
        file_path:    đường dẫn file .xlsx / .xlsm
        sheet_filter: chỉ đọc các sheet này (None = đọc tất cả)

    Returns:
        {
            "success":    bool,
            "structured": list[{"name": str, "rows": list[{"row": int, "cells": list}]}],
            "text_repr":  str  — biểu diễn tọa độ cho LLM prompt,
            "sheets":     list[str],
            "format":     "excel_structured",
            "error":      str | None
        }
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheets_to_read = sheet_filter if sheet_filter else wb.sheetnames
        sheets_data: list = []
        text_parts: list = []

        for sheet_name in sheets_to_read:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' không tồn tại trong {file_path}")
                continue

            ws = wb[sheet_name]
            rows_data: list = []
            sheet_text: list = [f"\n=== Sheet: {sheet_name} ==="]
            row_count = 0

            for row in ws.iter_rows():
                cells = []
                for cell in row:
                    val = cell.value
                    if val is not None and str(val).strip():
                        cells.append({
                            "col":        cell.column,
                            "col_letter": cell.column_letter,
                            "value":      str(val).strip(),
                        })

                if not cells:
                    continue

                row_num = row[0].row
                rows_data.append({"row": row_num, "cells": cells})

                cell_strs = [f"{c['col_letter']}={c['value']}" for c in cells]
                sheet_text.append(f"Row{row_num:4d}: {' | '.join(cell_strs)}")

                row_count += 1
                if row_count >= MAX_ROWS_PER_SHEET:
                    sheet_text.append(f"  ... (cắt bớt, tổng {ws.max_row} rows)")
                    break

            if rows_data:
                sheets_data.append({"name": sheet_name, "rows": rows_data})
                text_parts.extend(sheet_text)

        if not sheets_data:
            return {
                "success": False, "structured": [], "text_repr": "",
                "sheets": list(wb.sheetnames), "format": "excel_structured",
                "error": "File Excel không có dữ liệu",
            }

        return {
            "success":    True,
            "structured": sheets_data,
            "text_repr":  "\n".join(text_parts).strip(),
            "sheets":     list(wb.sheetnames),
            "format":     "excel_structured",
            "error":      None,
        }

    except ImportError:
        return _pandas_structured_fallback(file_path, sheet_filter)
    except Exception as e:
        logger.error(f"Lỗi đọc Excel structured {file_path}: {e}")
        return {
            "success": False, "structured": [], "text_repr": "",
            "sheets": [], "format": "excel_structured", "error": str(e),
        }


def _pandas_structured_fallback(file_path: str, sheet_filter=None) -> dict:
    """Fallback khi không có openpyxl: dùng pandas, mô phỏng row/col."""
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None, dtype=str, header=None)
        sheets_data, text_parts = [], []

        for sheet_name, df in all_sheets.items():
            if sheet_filter and sheet_name not in sheet_filter:
                continue
            df = df.fillna("").astype(str)
            rows_data, sheet_text = [], [f"\n=== Sheet: {sheet_name} ==="]

            for i, (_, row) in enumerate(df.iterrows()):
                cells = []
                for j, val in enumerate(row):
                    if val.strip():
                        letter = _col_letter(j + 1)
                        cells.append({"col": j + 1, "col_letter": letter, "value": val.strip()})
                if cells:
                    row_num = i + 1
                    rows_data.append({"row": row_num, "cells": cells})
                    cell_strs = [f"{c['col_letter']}={c['value']}" for c in cells]
                    sheet_text.append(f"Row{row_num:4d}: {' | '.join(cell_strs)}")
                if i >= MAX_ROWS_PER_SHEET:
                    break

            if rows_data:
                sheets_data.append({"name": sheet_name, "rows": rows_data})
                text_parts.extend(sheet_text)

        return {
            "success": True, "structured": sheets_data,
            "text_repr": "\n".join(text_parts).strip(),
            "sheets": list(all_sheets.keys()), "format": "excel_structured", "error": None,
        }
    except Exception as e:
        return {"success": False, "structured": [], "text_repr": "", "sheets": [],
                "format": "excel_structured", "error": str(e)}


def _col_letter(n: int) -> str:
    """Số cột (1-based) → chữ cái Excel: 1→A, 27→AA."""
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def read_excel(file_path: str) -> dict:
    """
    Đọc toàn bộ nội dung file Excel, convert về text thuần.
    Dùng cho backward compat — LLM prompt đơn giản.
    """
    try:
        all_sheets = pd.read_excel(file_path, sheet_name=None, dtype=str)
        logger.info(f"Đọc Excel: {file_path} ({len(all_sheets)} sheet)")

        full_text = []
        sheet_names = list(all_sheets.keys())

        for sheet_name, df in all_sheets.items():
            if df.empty:
                continue
            full_text.append(f"=== Sheet: {sheet_name} ===")
            df = df.dropna(how="all").dropna(axis=1, how="all").fillna("")
            full_text.append(df.to_string(index=False))
            full_text.append("")

        result_text = "\n".join(full_text).strip()

        if not result_text:
            return {"success": False, "text": "", "format": "excel",
                    "sheets": sheet_names, "error": "File Excel không có dữ liệu"}

        return {"success": True, "text": result_text, "format": "excel",
                "sheets": sheet_names, "error": None}

    except FileNotFoundError:
        return {"success": False, "text": "", "format": "excel",
                "sheets": [], "error": f"Không tìm thấy file: {file_path}"}
    except Exception as e:
        logger.error(f"Lỗi đọc Excel {file_path}: {e}")
        return {"success": False, "text": "", "format": "excel",
                "sheets": [], "error": f"Lỗi khi đọc Excel: {str(e)}"}


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else "sample_data/test_po.xlsx"
    mode = sys.argv[2] if len(sys.argv) > 2 else "structured"
    if mode == "structured":
        r = read_excel_structured(path)
        print(r["text_repr"] if r["success"] else r["error"])
    else:
        r = read_excel(path)
        print(r["text"][:500] if r["success"] else r["error"])
