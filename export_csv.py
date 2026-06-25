import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_test_excel(output_path="sample_data/processed_po.xlsx"):
    # 1. Giả lập dữ liệu đã được bóc tách từ đống Text Form Fields của bạn
    po_info = {
        "PO #:": "123456/22",
        "Date:": "22ND SEPTEMBER, 2022",
        "Client Name:": "FASHION ITEMS INC",
        "Total Amount:": "3,604.35$"
    }
    
    products = [
        {"Product Name": "Poshmark black dress", "Item #": "99880052", "Price": "1,080.00$", "Qty": 1, "Size": "M", "Colour": "BLACK", "Total": "1,080.00$"},
        {"Product Name": "Waterproof French overcoat", "Item #": "99881052", "Price": "995.00$", "Qty": 1, "Size": "M", "Colour": "BLACK", "Total": "995.00$"},
        {"Product Name": "Signature perfume", "Item #": "99885688", "Price": "205.00$", "Flv": 2, "Size": "*", "Colour": "*", "Total": "410.00$"},
        {"Product Name": "Classic blazer \"mindsweeper\"", "Item #": "99884899", "Price": "1,080.00$", "Qty": 1, "Size": "L", "Colour": "RED/BLACK", "Total": "1,080.00$"}
    ]

    # 2. Dùng openpyxl để tạo file Excel có format đẹp mắt
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Detail"
    
    # Bật hiển thị đường lưới gridlines
    ws.views.sheetView[0].showGridLines = True

    # Định dạng font và màu sắc
    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=11, bold=True)
    font_normal = Font(name="Arial", size=11)
    
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    # --- Phần 1: Tiêu đề & Thông tin chung ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "THÔNG TIN ĐƠN HÀNG (PURCHASE ORDER)"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = Alignment(horizontal="center")
    
    row_idx = 3
    for key, val in po_info.items():
        ws.cell(row=row_idx, column=1, value=key).font = font_bold
        ws.cell(row=row_idx, column=2, value=val).font = font_normal
        row_idx += 1

    # --- Phần 2: Danh sách sản phẩm ---
    row_idx += 1  # Cách ra 1 dòng
    headers = ["Product Name", "Item #", "Price", "Qty", "Size", "Colour", "Total"]
    
    # Ghi Header của bảng sản phẩm
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        
    # Ghi dữ liệu sản phẩm
    for prod in products:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=prod.get("Product Name")).font = font_normal
        ws.cell(row=row_idx, column=2, value=prod.get("Item #")).font = font_normal
        ws.cell(row=row_idx, column=3, value=prod.get("Price")).font = font_normal
        ws.cell(row=row_idx, column=4, value=prod.get("Qty") or prod.get("Flv")).font = font_normal # Đọc cả trường đặc biệt
        ws.cell(row=row_idx, column=5, value=prod.get("Size")).font = font_normal
        ws.cell(row=row_idx, column=6, value=prod.get("Colour")).font = font_normal
        ws.cell(row=row_idx, column=7, value=prod.get("Total")).font = font_normal

    # Tự động căn rộng các cột cho vừa chữ
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        
        # SỬA DÒNG NÀY: Dùng hàm get_column_letter bọc ngoài vị trí index của cột
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col[0].column)
        
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
    # Lưu file
    wb.save(output_path)
    print(f"File Excel test đã được tạo thành công tại: {output_path}")

if __name__ == "__main__":
    create_test_excel()