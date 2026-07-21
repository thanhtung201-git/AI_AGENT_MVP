"""
xlsx_pdf.py — Chuyển file Trimlist Excel sang PDF (khổ ngang A4).

Tách từ logic trong agent.py để các route khác (Task B...) tái sử dụng mà không
phụ thuộc vào endpoint agent. Định dạng Excel giả định: có một hàng header chứa
"NO" / "TRIM ITEM", các dòng meta nằm phía trên header.
"""
import os
from datetime import datetime

import openpyxl
from fpdf import FPDF

# Font Arial TTF để hỗ trợ tiếng Việt (Windows)
_ARIAL   = r"C:\Windows\Fonts\arial.ttf"
_ARIAL_B = r"C:\Windows\Fonts\arialbd.ttf"
_ARIAL_I = r"C:\Windows\Fonts\ariali.ttf"


def xlsx_to_pdf(xlsx_path: str, out_pdf_path: str, title: str = "TRIM LIST") -> str:
    """Đọc file Trimlist Excel và xuất ra PDF tại out_pdf_path. Trả về out_pdf_path."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Tìm header row
    header_idx = next(
        (i for i, r in enumerate(rows) if r and any(
            str(c or "").strip().upper() in ("NO", "TRIM ITEM", "TRIM_ITEM") for c in r
        )), None
    )

    # Meta (các dòng trước header)
    meta_lines = []
    if header_idx:
        for r in rows[:header_idx]:
            line = " | ".join(str(c or "").strip() for c in r if c)
            if line:
                meta_lines.append(line)

    # Data rows
    headers, data_rows = [], []
    if header_idx is not None:
        headers = [str(c or "").strip() for c in rows[header_idx]]
        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            cells = [str(c) if c is not None else "" for c in row]
            if cells and cells[0].strip().upper() in ("TOTAL", "TỔNG CỘNG", "GRAND TOTAL"):
                continue
            data_rows.append(cells)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("Arial", "",  _ARIAL)
    pdf.add_font("Arial", "B", _ARIAL_B)
    pdf.add_font("Arial", "I", _ARIAL_I)
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_margins(10, 10, 10)

    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, title, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Arial", "", 8)
    for line in meta_lines[:5]:
        pdf.cell(0, 5, line[:120], align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    if headers and data_rows:
        page_w = pdf.w - 20
        col_widths = [max(len(h), max((len(str(r[i])) if i < len(r) else 0 for r in data_rows), default=0))
                      for i, h in enumerate(headers)]
        total_chars = sum(col_widths) or 1
        col_w = [max(12, (page_w * w / total_chars)) for w in col_widths]
        actual_total = sum(col_w)
        if actual_total > page_w:
            col_w = [w * page_w / actual_total for w in col_w]

        pdf.set_font("Arial", "B", 7)
        pdf.set_fill_color(63, 84, 186)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 6, h[:30], border=1, fill=True)
        pdf.ln()

        pdf.set_font("Arial", "", 7)
        pdf.set_text_color(0, 0, 0)
        for ri, row in enumerate(data_rows):
            fill = (ri % 2 == 0)
            pdf.set_fill_color(245, 246, 252) if fill else pdf.set_fill_color(255, 255, 255)
            for i in range(len(headers)):
                val = row[i] if i < len(row) else ""
                pdf.cell(col_w[i], 5.5, str(val)[:50], border=1, fill=fill)
            pdf.ln()
    else:
        pdf.set_font("Arial", "I", 9)
        pdf.cell(0, 8, "Không có dữ liệu", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Arial", "I", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"MCNA Garment — AI Agent | Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
             align="R", new_x="LMARGIN", new_y="NEXT")

    os.makedirs(os.path.dirname(out_pdf_path), exist_ok=True)
    pdf.output(out_pdf_path)
    return out_pdf_path
