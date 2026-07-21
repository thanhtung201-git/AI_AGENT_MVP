"""
update_file_extractor.py — Orchestrator hybrid cho file update từ buyer.

Flow:
  1. Rule-based   → HZSHExtractor (nhanh, không tốn token)
  2. LLM fallback → Groq (linh hoạt, xử lý format lạ)
  3. Manual       → trả flag needs_manual_input=True cho frontend

Hỗ trợ: .xlsx, .xls, .pdf (qua tools.reader)
"""
import os
import json
import logging
import re
from typing import Dict, Any

logger = logging.getLogger(__name__)

_LLM_SYSTEM_PROMPT = """Bạn là AI chuyên đọc file update số lượng hàng may mặc từ buyer.

Nhiệm vụ: Tìm bảng màu/size/số lượng trong nội dung file và trả về JSON chuẩn.

OUTPUT FORMAT (chỉ JSON, không thêm text):
{
  "style_code": "<mã style hoặc null>",
  "sizes": ["S", "M", "L", "XL"],
  "colors": [
    {"code": "N2", "name": "NORMAL NAVY", "qty_per_size": {"S": 100, "M": 200}, "total": 300}
  ],
  "total_qty": 1538,
  "size_breakdown": {"S": 351, "M": 589, "L": 422, "XL": 176},
  "confidence": 0.9
}

Lưu ý:
- Nếu không có mã màu, để code = tên màu
- Nếu chỉ có tổng qty, không có breakdown: sizes=[], colors có total nhưng qty_per_size={}
- confidence: 0.0-1.0 (độ chắc chắn của bạn về kết quả)
- Nếu không tìm thấy bảng qty nào: trả {"error": "no_qty_table", "confidence": 0}
"""


def _strip_think(text: str) -> str:
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
    text = re.sub(r"^```(?:json)?\s*", "", text.strip())
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


class UpdateFileExtractor:
    """
    Hybrid extractor cho file update từ buyer.
    Thử rule-based trước, fallback LLM nếu fail.
    """

    def extract(self, file_path: str, filename: str = "") -> Dict[str, Any]:
        """
        Returns:
        {
          "success": bool,
          "method": "rule_based" | "llm" | "manual",
          "needs_manual_input": bool,   # True nếu cả 2 đều fail
          "file_type": "HZSH" | "QTY_UPDATE" | "UNKNOWN",
          "style_code": str | None,
          "sizes": [...],
          "colors": [...],
          "total_qty": int,
          "size_breakdown": {...},
          "confidence": float,
          "raw_file": str,
          "error": str | None,
        }
        """
        if not filename:
            filename = os.path.basename(file_path)

        base = {
            "raw_file": filename,
            "style_code": None,
            "sizes": [],
            "colors": [],
            "total_qty": 0,
            "size_breakdown": {},
            "confidence": 0.0,
        }

        # ── Bước 1: Rule-based ────────────────────────────────────────────────
        rule_result = self._try_rule_based(file_path, filename)

        if rule_result.get("success"):
            quality = self._quality_score(rule_result)
            if quality >= 0.8:
                # Tốt — dùng luôn
                logger.info(f"UpdateFileExtractor: rule-based OK (quality={quality:.1f}) cho {filename}")
                return {**base, **rule_result, "method": "rule_based", "needs_manual_input": False}
            else:
                # Có dữ liệu nhưng chất lượng thấp — chạy thêm LLM để bổ sung
                logger.info(
                    f"UpdateFileExtractor: rule-based partial (quality={quality:.1f}) — "
                    f"thử LLM để bổ sung ({self._quality_issues(rule_result)})"
                )
                llm_result = self._try_llm(file_path, filename)
                if llm_result.get("success"):
                    merged = self._merge(rule_result, llm_result)
                    logger.info("UpdateFileExtractor: merged rule+LLM OK")
                    return {**base, **merged, "method": "rule_based+llm", "needs_manual_input": False}
                # LLM fail → dùng rule-based dù chất lượng thấp (vẫn hơn không có gì)
                logger.warning("UpdateFileExtractor: LLM bổ sung fail — dùng rule-based partial")
                return {**base, **rule_result, "method": "rule_based_partial", "needs_manual_input": False}

        logger.info(f"UpdateFileExtractor: rule-based fail ({rule_result.get('error')}) — thử LLM")

        # ── Bước 2: LLM fallback ──────────────────────────────────────────────
        llm_result = self._try_llm(file_path, filename)
        if llm_result.get("success"):
            logger.info(f"UpdateFileExtractor: LLM OK cho {filename}")
            return {**base, **llm_result, "method": "llm", "needs_manual_input": False}

        logger.warning(f"UpdateFileExtractor: LLM fail ({llm_result.get('error')}) — yêu cầu manual")

        # ── Bước 3: Manual input ──────────────────────────────────────────────
        return {
            **base,
            "success":            False,
            "method":             "manual",
            "needs_manual_input": True,
            "file_type":          "UNKNOWN",
            "error": (
                f"Không tự động đọc được file này. "
                f"Rule-based: {rule_result.get('error')}. "
                f"LLM: {llm_result.get('error')}."
            ),
        }

    # ── Quality check ─────────────────────────────────────────────────────────

    def _quality_score(self, result: Dict[str, Any]) -> float:
        """
        Chấm điểm chất lượng kết quả rule-based: 0.0 → 1.0.
        Dùng để quyết định có cần chạy thêm LLM không.
        """
        score = 0.0
        colors = result.get("colors") or []
        sizes  = result.get("sizes")  or []
        total  = result.get("total_qty") or 0

        # Có colors → +0.4
        if colors:
            score += 0.4

        # Có sizes → +0.3
        if sizes:
            score += 0.3

        # total_qty > 0 → +0.2
        if total > 0:
            score += 0.2

        # Tổng color.total == total_qty (nhất quán) → +0.1
        color_sum = sum(c.get("total", 0) for c in colors)
        if total > 0 and color_sum > 0 and abs(color_sum - total) <= 1:
            score += 0.1

        return round(score, 2)

    def _quality_issues(self, result: Dict[str, Any]) -> str:
        """Trả về mô tả ngắn về những gì còn thiếu."""
        issues = []
        if not result.get("colors"):
            issues.append("không có colors")
        if not result.get("sizes"):
            issues.append("không có sizes")
        if not result.get("total_qty"):
            issues.append("total_qty=0")
        colors = result.get("colors") or []
        total  = result.get("total_qty") or 0
        color_sum = sum(c.get("total", 0) for c in colors)
        if total > 0 and color_sum > 0 and abs(color_sum - total) > 1:
            issues.append(f"sum mismatch ({color_sum} vs {total})")
        return ", ".join(issues) if issues else "OK"

    # ── Merge rule + LLM ──────────────────────────────────────────────────────

    def _merge(self, rule: Dict[str, Any], llm: Dict[str, Any]) -> Dict[str, Any]:
        """
        Kết hợp rule-based và LLM: lấy rule-based làm nền,
        dùng LLM để điền vào những chỗ còn trống/bất thường.
        """
        merged = dict(rule)  # bắt đầu từ rule-based

        # Sizes: nếu rule không có size → dùng LLM
        if not merged.get("sizes") and llm.get("sizes"):
            merged["sizes"] = llm["sizes"]
            logger.info("merge: lấy sizes từ LLM")

        # Colors: nếu rule không có colors → dùng LLM
        if not merged.get("colors") and llm.get("colors"):
            merged["colors"] = llm["colors"]
            logger.info("merge: lấy colors từ LLM")

        # Nếu rule có colors nhưng LLM có nhiều hơn (bắt thêm được) → dùng LLM
        if len(llm.get("colors") or []) > len(merged.get("colors") or []):
            merged["colors"] = llm["colors"]
            logger.info("merge: LLM có nhiều colors hơn rule-based")

        # size_breakdown: nếu rule trống → dùng LLM
        if not merged.get("size_breakdown") and llm.get("size_breakdown"):
            merged["size_breakdown"] = llm["size_breakdown"]

        # total_qty: ưu tiên giá trị > 0 và nhất quán với colors
        rule_total = rule.get("total_qty") or 0
        llm_total  = llm.get("total_qty")  or 0
        color_sum  = sum(c.get("total", 0) for c in (merged.get("colors") or []))

        if rule_total == 0 and llm_total > 0:
            merged["total_qty"] = llm_total
        elif color_sum > 0 and abs(color_sum - rule_total) > 5:
            # LLM total gần color_sum hơn → dùng LLM
            if abs(color_sum - llm_total) < abs(color_sum - rule_total):
                merged["total_qty"] = llm_total
                logger.info(f"merge: dùng total_qty từ LLM ({llm_total}) vì gần color_sum hơn")

        # Confidence = trung bình có trọng số
        merged["confidence"] = round(
            rule.get("confidence", 0.5) * 0.6 + llm.get("confidence", 0.5) * 0.4, 2
        )
        merged["success"] = True
        return merged

    # ── Rule-based — nhận diện theo NỘI DUNG, không theo tên file ───────────

    def _try_rule_based(self, file_path: str, filename: str) -> Dict[str, Any]:
        ext = os.path.splitext(file_path)[-1].lower()

        # Chỉ áp dụng rule-based cho Excel
        if ext not in (".xlsx", ".xls"):
            return {"success": False, "error": "Không phải Excel — để LLM xử lý"}

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Không mở được file: {e}"}

        detected = self._detect_format(wb)
        wb.close()

        if detected == "COLOR_SIZE_ASSORTMENT":
            # Bảng color × size (HZSH và các buyer khác cùng cấu trúc)
            try:
                from backend.extractors.hzsh_extractor import HZSHExtractor
                result = HZSHExtractor().extract(file_path)
                if result.get("success"):
                    return {**result, "file_type": "COLOR_SIZE_ASSORTMENT", "confidence": 1.0}
                return {"success": False, "error": result.get("error")}
            except Exception as e:
                return {"success": False, "error": str(e)}

        # Thêm rule cho format mới khi phát hiện:
        # if detected == "QTY_BREAKDOWN_V2":
        #     return NewFormatExtractor().extract(file_path)

        return {"success": False, "error": f"Không nhận diện được cấu trúc file (format={detected}) — chuyển sang LLM"}

    def _detect_format(self, wb) -> str:
        """
        Nhận diện cấu trúc file bằng cách scan 30 dòng đầu mỗi sheet.
        Trả về tên format hoặc 'UNKNOWN'.

        Để thêm format mới: thêm điều kiện vào đây + extractor tương ứng trong _try_rule_based.
        """
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=30, values_only=True):
                joined = " ".join(str(c or "").strip().upper() for c in row if c)

                # Bảng color × size — HZSH (Hazzys) và buyer khác có cùng cấu trúc
                if "COLOR" in joined and "SIZE" in joined and "ASSORTMENT" in joined:
                    return "COLOR_SIZE_ASSORTMENT"

                # Placeholder cho format khác — thêm vào khi gặp file mới:
                # if "QTY BREAKDOWN" in joined and "COLOR" in joined:
                #     return "QTY_BREAKDOWN_V2"
                # if "PURCHASE ORDER" in joined and "SIZE RATIO" in joined:
                #     return "PO_SIZE_RATIO"

        return "UNKNOWN"

    # ── LLM fallback ──────────────────────────────────────────────────────────

    def _try_llm(self, file_path: str, filename: str) -> Dict[str, Any]:
        # Đọc nội dung file
        try:
            content = self._read_file_content(file_path)
        except Exception as e:
            return {"success": False, "error": f"Không đọc được file: {e}"}

        if not content or len(content.strip()) < 20:
            return {"success": False, "error": "File rỗng hoặc không có text"}

        # Giới hạn token
        content_trimmed = content[:6000]

        try:
            from backend.utils.groq_client import GroqClient
            client = GroqClient()
            raw = client.extract_json(
                system_prompt=_LLM_SYSTEM_PROMPT,
                user_content=(
                    f"Tên file: {filename}\n\n"
                    f"Nội dung:\n{content_trimmed}"
                ),
            )
        except Exception as e:
            return {"success": False, "error": f"LLM lỗi: {e}"}

        # Validate response
        if not isinstance(raw, dict):
            return {"success": False, "error": "LLM không trả JSON hợp lệ"}

        if raw.get("error") == "no_qty_table" or not raw.get("colors"):
            return {"success": False, "error": "LLM không tìm thấy bảng số lượng"}

        confidence = float(raw.get("confidence", 0.5))
        if confidence < 0.3:
            return {"success": False, "error": f"LLM không chắc chắn (confidence={confidence:.1f})"}

        colors = raw.get("colors", [])
        total_qty = raw.get("total_qty") or sum(c.get("total", 0) for c in colors)

        return {
            "success":        True,
            "file_type":      "QTY_UPDATE",
            "style_code":     raw.get("style_code"),
            "sizes":          raw.get("sizes", []),
            "colors":         colors,
            "total_qty":      total_qty,
            "size_breakdown": raw.get("size_breakdown", {}),
            "confidence":     confidence,
        }

    def _read_file_content(self, file_path: str) -> str:
        ext = os.path.splitext(file_path)[-1].lower()

        if ext in (".xlsx", ".xls"):
            return self._read_excel_as_text(file_path)

        # PDF, docx... dùng tools.reader
        try:
            from tools.reader import read_file
            result = read_file(file_path)
            if result.get("success"):
                return result.get("text", "")
        except Exception:
            pass
        return ""

    def _read_excel_as_text(self, file_path: str) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append(" | ".join(cells))
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            logger.warning(f"_read_excel_as_text error: {e}")
            return ""
