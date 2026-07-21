"""
buying_sheet_extractor.py — Rule-based extractor cho format "buying sheet" của Tessellation/TSSL.

Cấu trúc file:
  Row 4: DS | 발주 | SMS | BULK | NO | DES./COLOR CMNT | Color Code | TTL | KOREA | ... | TAIWAN | ... | VIETNAM | ... | INDIA | ...
  Row 5: sub-size headers (90/0XS, 95/00S, 100/00M, 105/00L, 110/0XL, 115/XXL...)
  Data : BULK=style (new group) + Color=TTL (tổng) | Color=WT/B2/G2/K2... (màu cụ thể)

Detect: header row có "Color" + "TTL" và subrow có size codes số (90-200).
"""
import logging
import re
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _int(v) -> int:
    try:
        return int(v or 0)
    except (ValueError, TypeError):
        return 0


def _size_code(raw) -> Optional[str]:
    """Lấy size code số từ cell như '90\\n0XS' → '90', hoặc 'S', 'M'..."""
    if raw is None:
        return None
    s = str(raw).strip().split("\n")[0].strip().upper()
    # Chỉ lấy dạng số (90-200) hoặc chữ S/M/L/XL/XXL...
    if re.match(r"^\d{2,3}$", s) and 60 <= int(s) <= 200:
        return s
    if s in ("XS", "S", "M", "L", "XL", "XXL", "2XL", "3XL"):
        return s
    return None


class BuyingSheetExtractor:
    """
    Đọc buying sheet (Tessellation/TSSL format) và trả về danh sách styles
    với màu sắc, size breakdown, và delivery date.
    """

    # Cột cố định (1-indexed, khớp với file HZ)
    _BULK_COL  = 4   # style code
    _COLOR_COL = 7   # color code / "TTL"
    _TTL_COL   = 8   # total qty
    _DEL_COL   = 34  # delivery date

    @classmethod
    def detect(cls, file_path: str) -> bool:
        """Trả True nếu file là buying sheet format."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active
            for i in range(1, 8):
                row_vals = [str(ws.cell(row=i, column=j).value or "").upper()
                            for j in range(1, 15)]
                joined = " ".join(row_vals)
                if "COLOR" in joined and "TTL" in joined and "BULK" in joined:
                    wb.close()
                    return True
            wb.close()
        except Exception:
            pass
        return False

    @classmethod
    def extract_all(cls, file_path: str) -> List[Dict[str, Any]]:
        """
        Trả về list các style dicts:
        {
          "style_code": "HZSH6F201",
          "description": "에센셜 옥스포드 셔츠",
          "delivery_date": "2026-06-26",
          "total_qty": 2104,
          "sizes": ["90", "95", "100", "105", "110", "115"],
          "colors": [
            {"code": "WT", "name": "WT", "qty_per_size": {"90":115, "95":187, ...}, "total": 596},
            ...
          ],
          "size_breakdown": {"90": 115, "95": ...},
        }
        """
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # ── Tìm header row và size mapping ────────────────────────────────────
        header_row, size_col_map = cls._find_layout(ws)
        if header_row is None:
            logger.warning("BuyingSheetExtractor: không tìm thấy header row")
            wb.close()
            return []

        sizes_ordered = list(size_col_map.keys())

        # ── Đọc data rows ─────────────────────────────────────────────────────
        styles: List[Dict] = []
        current: Optional[Dict] = None

        for r in range(header_row + 2, ws.max_row + 1):
            bulk  = _cell(ws, r, cls._BULK_COL)
            color = _cell(ws, r, cls._COLOR_COL)
            ttl   = _int(_cell(ws, r, cls._TTL_COL))

            if bulk is None and color is None:
                continue  # blank row

            # Hàng TTL của style mới
            if bulk and str(color or "").strip().upper() == "TTL":
                desc     = str(_cell(ws, r, 6) or "").strip()
                del_raw  = _cell(ws, r, cls._DEL_COL)
                del_date = ""
                if hasattr(del_raw, "strftime"):
                    del_date = del_raw.strftime("%Y-%m-%d")

                current = {
                    "style_code":   str(bulk).strip(),
                    "description":  desc,
                    "delivery_date": del_date,
                    "total_qty":    ttl,
                    "sizes":        list(sizes_ordered),
                    "colors":       [],
                    "size_breakdown": {},
                }
                styles.append(current)
                continue

            # Hàng màu cụ thể (bulk trống, color = mã màu)
            if current is not None and color and str(color).strip().upper() not in ("TTL", "COLOR\nCODE", "COLOR CODE"):
                color_code = str(color).strip().upper()
                # Bỏ qua nếu là header row bị lọt (thường có chứa \n)
                if "\n" in color_code and "COLOR" in color_code:
                    continue
                qty_by_size: Dict[str, int] = {}
                for sz, cols in size_col_map.items():
                    q = sum(_int(_cell(ws, r, c)) for c in cols)
                    if q > 0:
                        qty_by_size[sz] = q

                color_total = sum(qty_by_size.values())
                if color_total == 0:
                    color_total = ttl  # fallback TTL column

                current["colors"].append({
                    "code":         color_code,
                    "name":         color_code,
                    "qty_per_size": qty_by_size,
                    "total":        color_total,
                })

        # ── Tính size_breakdown cho mỗi style ────────────────────────────────
        for s in styles:
            bd: Dict[str, int] = {}
            for sz in s["sizes"]:
                bd[sz] = sum(c["qty_per_size"].get(sz, 0) for c in s["colors"])
            s["size_breakdown"] = {k: v for k, v in bd.items() if v > 0}
            # Giữ sizes chỉ có data
            s["sizes"] = [sz for sz in s["sizes"] if bd.get(sz, 0) > 0]

        wb.close()
        logger.info(f"BuyingSheetExtractor: tìm thấy {len(styles)} styles")
        return styles

    @classmethod
    def _find_layout(cls, ws) -> Tuple[Optional[int], Dict[str, List[int]]]:
        """
        Tìm header row và map size_code → [col_indices].
        Trả về (header_row, {size_code: [col1, col2, ...]}).
        """
        header_row = None
        size_header_row = None

        # Tìm header row bằng "Color" + "TTL" + "BULK"
        for i in range(1, min(10, ws.max_row + 1)):
            vals = [str(ws.cell(row=i, column=j).value or "").strip().upper()
                    for j in range(1, min(15, ws.max_column + 1))]
            joined = " ".join(vals)
            if "COLOR" in joined and "TTL" in joined and "BULK" in joined:
                header_row = i
                size_header_row = i + 1
                break

        if header_row is None:
            return None, {}

        # Đọc size header row — map size_code → list of columns
        size_col_map: Dict[str, List[int]] = {}
        for j in range(1, ws.max_column + 1):
            sz = _size_code(ws.cell(row=size_header_row, column=j).value)
            if sz:
                size_col_map.setdefault(sz, []).append(j)

        # Sắp xếp sizes theo thứ tự số tăng dần
        def _sort_key(s: str) -> int:
            try:
                return int(s)
            except ValueError:
                return {"XS": 0, "S": 1, "M": 2, "L": 3, "XL": 4, "XXL": 5, "2XL": 6}.get(s, 99)

        size_col_map = dict(sorted(size_col_map.items(), key=lambda x: _sort_key(x[0])))

        logger.debug(f"BuyingSheetExtractor: header_row={header_row}, sizes={list(size_col_map.keys())}")
        return header_row, size_col_map

    @classmethod
    def extract_style(cls, file_path: str, style_code: str) -> Optional[Dict[str, Any]]:
        """Trích xuất 1 style cụ thể. Trả None nếu không tìm thấy."""
        all_styles = cls.extract_all(file_path)
        code_upper = style_code.strip().upper()
        for s in all_styles:
            if s["style_code"].upper() == code_upper:
                return s
        return None
