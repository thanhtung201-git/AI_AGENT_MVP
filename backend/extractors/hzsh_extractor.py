"""
hzsh_extractor.py — Đọc file HZSH Excel (Color/Size Assortment update từ Hazzys buyer).

Cấu trúc file:
  - Dòng "COLOR/SIZE ASSORTMENT"
  - Header row: Color | Size | [label...] | 00S | 00M | 00L | 0XL | XXL | ... | SUM
  - Data rows:  N2 | NORMAL NAVY | ... | 191 | 283 | 206 | 73 | 21 | ... | 774
  - SUM row:    SUM | | ... | 351 | 589 | 422 | 134 | 42 | ... | 1538

Không dùng LLM — parse trực tiếp bằng openpyxl.
"""
import os
import re
import logging
from typing import Dict, Any, List, Optional

logger = logging.getLogger(__name__)

# Size pattern: 00S, 00M, 00L, 0XL, XXL, XS, S, M, L, XL, 2XL, 3XL...
_SIZE_RE = re.compile(
    r"^(0{0,2}(XS|S|M|L|XL|XXL|2XL|3XL|4XL)|[0-9X]+[SML]|FREE SIZE|FREE)$",
    re.IGNORECASE,
)


def _is_size(val: str) -> bool:
    return bool(_SIZE_RE.match(val.strip()))


def _to_int(val) -> int:
    if val is None:
        return 0
    s = str(val).replace(",", "").replace(".", "").strip()
    if s in ("-", "", "None"):
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


class HZSHExtractor:
    """Trích xuất thông tin size/qty update từ file HZSH Excel."""

    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Returns:
        {
          "success": bool,
          "style_code": "6C33101",         # từ tên file HZSH6C33101
          "sizes": ["00S","00M","00L","0XL","XXL"],
          "colors": [
            {"code":"N2","name":"NORMAL NAVY","qty_per_size":{"00S":191,...},"total":774},
            ...
          ],
          "total_qty": 1538,
          "size_breakdown": {"00S":351,"00M":589,...},
          "raw_file": "HZSH6C33101.xlsx",
          "error": None
        }
        """
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "error": "openpyxl chưa được cài. Chạy: pip install openpyxl"}

        if not os.path.exists(file_path):
            return {"success": False, "error": f"File không tồn tại: {file_path}"}

        # Style code từ tên file: HZSH6C33101.xlsx → 6C33101
        basename   = os.path.basename(file_path)
        style_code = self._parse_style_from_filename(basename)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return {"success": False, "error": f"Không mở được file: {e}"}

        # Thử từng sheet — lấy sheet đầu tiên có "COLOR/SIZE ASSORTMENT"
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = self._parse_sheet(ws, style_code, basename)
            if result.get("success"):
                wb.close()
                return result

        wb.close()
        return {
            "success":    False,
            "style_code": style_code,
            "raw_file":   basename,
            "error":      "Không tìm thấy bảng COLOR/SIZE ASSORTMENT trong file",
        }

    def _parse_style_from_filename(self, filename: str) -> str:
        """HZSH6C33101.xlsx → 6C33101"""
        name = os.path.splitext(filename)[0]
        m = re.match(r"HZSH(.+)", name, re.IGNORECASE)
        if m:
            return m.group(1).strip()
        return name

    def _parse_sheet(self, ws, style_code: str, raw_file: str) -> Dict[str, Any]:
        rows = list(ws.iter_rows(values_only=True))

        # 1. Tìm dòng "COLOR/SIZE ASSORTMENT"
        assortment_row = None
        for i, row in enumerate(rows):
            joined = " ".join(str(c or "").strip().upper() for c in row if c)
            if "COLOR" in joined and "SIZE" in joined and "ASSORTMENT" in joined:
                assortment_row = i
                break

        if assortment_row is None:
            return {"success": False}

        # 2. Tìm header row (chứa cột kích cỡ như 00S, 00M...)
        header_row_idx   = None
        size_col_indices: List[int] = []
        sizes_list:       List[str] = []
        sum_col_idx: Optional[int]  = None

        for i in range(assortment_row + 1, min(assortment_row + 10, len(rows))):
            row = rows[i]
            sizes_found = []
            tmp_sum = None
            for j, cell in enumerate(row):
                val = str(cell or "").strip().upper()
                if _is_size(val):
                    sizes_found.append((j, val))
                if val == "SUM":
                    tmp_sum = j

            if len(sizes_found) >= 2:
                header_row_idx   = i
                size_col_indices = [j for j, _ in sizes_found]
                sizes_list       = [v for _, v in sizes_found]
                sum_col_idx      = tmp_sum
                break

        if header_row_idx is None or not size_col_indices:
            return {"success": False}

        # Fallback tìm SUM col từ header row nếu chưa có
        if sum_col_idx is None:
            for j, cell in enumerate(rows[header_row_idx]):
                if str(cell or "").strip().upper() == "SUM":
                    sum_col_idx = j
                    break

        # 3. Parse data rows
        colors: List[Dict] = []
        size_totals = {s: 0 for s in sizes_list}

        for i in range(header_row_idx + 1, len(rows)):
            row = rows[i]
            if not any(row):
                continue

            first_vals = [str(c or "").strip().upper() for c in row[:4] if c]

            # Dòng TOTAL SUM → dừng
            if any(v in ("SUM", "TOTAL SUM", "TOTAL") for v in first_vals[:2]):
                break

            # Dòng COLOR SUM → bỏ qua
            if any(v == "COLOR SUM" for v in first_vals):
                continue

            color_code = str(row[0] or "").strip() if len(row) > 0 else ""
            color_name = str(row[1] or "").strip() if len(row) > 1 else ""

            if not color_code or color_code.upper() in ("COLOR", ""):
                continue

            qty_per_size: Dict[str, int] = {}
            for idx, size in zip(size_col_indices, sizes_list):
                qty_per_size[size] = _to_int(row[idx] if idx < len(row) else None)

            row_total = _to_int(row[sum_col_idx] if sum_col_idx is not None and sum_col_idx < len(row) else None)
            if row_total == 0:
                row_total = sum(qty_per_size.values())

            if row_total == 0 and not color_name:
                continue

            colors.append({
                "code":         color_code,
                "name":         color_name,
                "qty_per_size": qty_per_size,
                "total":        row_total,
            })
            for size in sizes_list:
                size_totals[size] += qty_per_size.get(size, 0)

        total_qty = sum(c["total"] for c in colors)

        if not colors:
            return {"success": False}

        return {
            "success":        True,
            "style_code":     style_code,
            "sizes":          sizes_list,
            "colors":         colors,
            "total_qty":      total_qty,
            "size_breakdown": size_totals,
            "raw_file":       raw_file,
            "error":          None,
        }
