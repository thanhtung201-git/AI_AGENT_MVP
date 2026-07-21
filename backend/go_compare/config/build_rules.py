"""
build_rules.py — Convert the company's "Rules create Batch GO.xlsx" into the JSON
config the filler reads at runtime. Run this whenever the rules workbook changes:

    python -m backend.go_compare.config.build_rules

Output (same folder):
  batch_go_rules.json   — per-field rule: how each cell is filled, plus the buyer's
                          fixed value for that field when the workbook carries one

The buyer column is headed "Hazzys (auto fill once end user input buyer name)":
those values are FIXED PER BUYER and the agent is meant to fill them in once the
end user names the buyer. They are exported here as `buyer_value`.

Which of them the agent may actually write is deliberately narrow — see
batch_go_buyer_fill.json. Only the fields listed there are auto-filled; every other
`end_user_fill` cell stays blank. Widen that list, not this script, to fill more.

Keeping the rules and the buyer values in JSON (not Python) honours the project
rule: no hardcoded customer data — only LLM + JSON config.
"""
import json
import os
import re

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
_SRC  = os.path.join(_HERE, "batch_go_rules_source.xlsx")
_SUMMARY_SHEET = "Tổng hợp"
# Column G: "<Buyer> (auto fill once end user input buyer name)" — the buyer's fixed
# value for that field. One buyer per workbook today; add columns for more.
_BUYER_COL = 7


def _rule_type(rule_text: str) -> str:
    f = (rule_text or "").strip().lower()
    if not f or "no need fill" in f:
        return "no_need_fill"
    if "always" in f:
        return "always"
    if "each go" in f:
        return "seq_go"
    if "each color" in f:
        return "seq_color"
    if "each size" in f:
        return "seq_size"
    if "each lot" in f:
        return "seq_lot"
    if "cust style desc" in f:
        return "style_desc"
    if "follow po" in f:
        return "follow_po"
    if "end user fill" in f:
        return "end_user_fill"
    return "end_user_fill"


def _always_value(rule_text: str, g_value) -> str:
    """Extract the constant for an `Always = "X"` rule; fall back to the buyer col."""
    m = re.search(r'always\s*=?\s*"([^"]*)"', rule_text or "", re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"always\s*=?\s*([A-Za-z0-9/_-]+)", rule_text or "", re.IGNORECASE)
    if m and m.group(1).lower() != "none":
        return m.group(1)
    if "none" in (rule_text or "").lower():
        return "None"
    return "" if g_value is None else str(g_value)


def build():
    wb = openpyxl.load_workbook(_SRC, data_only=True)
    ws = wb[_SUMMARY_SHEET]

    # Header row is row 2: A=STT B=Sheet C=Section D=Field E=Mandatory F=Rules
    rules = []
    for r in range(3, ws.max_row + 1):
        sheet   = ws.cell(r, 2).value
        section = ws.cell(r, 3).value
        field   = ws.cell(r, 4).value
        if not (sheet and section and field):
            continue
        mand = str(ws.cell(r, 5).value or "").strip()
        ftxt = str(ws.cell(r, 6).value or "").strip()
        rtype = _rule_type(ftxt)

        entry = {
            "sheet":     str(sheet).strip(),
            "section":   str(section).strip(),
            "field":     str(field).strip(),
            "mandatory": mand,
            "rule":      rtype,
        }
        if rtype == "always":
            entry["value"] = _always_value(ftxt, ws.cell(r, 7).value)
        buyer_val = ws.cell(r, _BUYER_COL).value
        if buyer_val not in (None, ""):
            entry["buyer_value"] = str(buyer_val).strip()
        rules.append(entry)

    with open(os.path.join(_HERE, "batch_go_rules.json"), "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=1)

    print(f"Wrote {len(rules)} field rules")


if __name__ == "__main__":
    build()
