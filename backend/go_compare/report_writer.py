"""
report_writer.py — Deterministic writers for the comparison outputs.

  - Compare_Report.xlsx : columns Status | PO Value | GO Value | Difference | Source | Confidence
  - Alerts.json         : ERROR / WARNING / INFO list

No LLM. Pure deterministic rendering of the compare result.
"""
import json
import logging
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HDR    = PatternFill("solid", fgColor="1F3864")
_WHITE  = Font(color="FFFFFF", bold=True, size=10)
_BOLD   = Font(bold=True, size=10)
_NORM   = Font(size=9)
_CEN    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN   = Border(*[Side(style="thin")] * 4)

_STATUS_FILL = {
    "MATCH":    PatternFill("solid", fgColor="E8F5E8"),
    "MISMATCH": PatternFill("solid", fgColor="FFD7D7"),
    "MISSING":  PatternFill("solid", fgColor="FFE0B3"),
    "EXTRA":    PatternFill("solid", fgColor="EEF4FF"),
}
_LEVEL_FILL = {
    "ERROR":   PatternFill("solid", fgColor="FFD7D7"),
    "WARNING": PatternFill("solid", fgColor="FFF3CC"),
    "INFO":    PatternFill("solid", fgColor="EEF4FF"),
}


class ReportWriter:

    def write_compare_report(
        self, compare: Dict[str, Any], output_path: str
    ) -> str:
        wb = Workbook()
        self._summary_sheet(wb, compare)
        self._detail_sheet(wb, compare.get("rows") or [])
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(output_path)
        logger.info(f"ReportWriter: saved {output_path}")
        return output_path

    def write_alerts_json(self, compare: Dict[str, Any], output_path: str) -> str:
        payload = {
            "summary": compare.get("summary") or {},
            "alerts":  compare.get("alerts") or [],
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        logger.info(f"ReportWriter: saved {output_path}")
        return output_path

    # ── sheets ────────────────────────────────────────────────────────────────

    def _summary_sheet(self, wb: Workbook, compare: Dict[str, Any]):
        ws = wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 22

        s = compare.get("summary") or {}
        status = s.get("status", "—")
        fill = {"OK": "E8F5E8", "PARTIAL": "FFF3CC", "MISMATCH": "FFD7D7"}.get(status, "F2F2F2")

        ws.merge_cells("A1:B1")
        c = ws.cell(1, 1, f"COMPARE REPORT — {status}")
        c.fill, c.font, c.alignment = PatternFill("solid", fgColor=fill), Font(bold=True, size=12), _CEN

        rows = [
            ("Errors",        s.get("errors", 0)),
            ("Warnings",      s.get("warnings", 0)),
            ("Info",          s.get("infos", 0)),
            ("PO Total Qty",  f"{s.get('po_total', 0):,.0f}"),
            ("GO Total Qty",  f"{s.get('go_total', 0):,.0f}"),
            ("Qty Difference", f"{s.get('qty_diff', 0):+,.0f}"),
            ("PO Lines",      s.get("po_lines", 0)),
            ("GO Lines",      s.get("go_lines", 0)),
            ("Rows Compared", s.get("compared", 0)),
        ]
        r = 2
        for label, val in rows:
            ws.cell(r, 1, label).font = _BOLD
            ws.cell(r, 1).border = _THIN
            ws.cell(r, 2, val).font = _NORM
            ws.cell(r, 2).border = _THIN
            ws.cell(r, 2).alignment = _CEN
            r += 1

    def _detail_sheet(self, wb: Workbook, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet("Detail", 1)
        ws.sheet_view.showGridLines = False

        headers = ["Field", "Status", "PO Value", "GO Value", "Difference", "Source", "Confidence"]
        widths  = [30, 12, 16, 16, 14, 42, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 26
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill, cell.font, cell.alignment, cell.border = _HDR, _WHITE, _CEN, _THIN

        r = 2
        for row in rows:
            status = row.get("status", "")
            sfill  = _STATUS_FILL.get(status, PatternFill("solid", fgColor="FFFFFF"))
            vals = [
                row.get("field", ""), status, row.get("po_value", ""),
                row.get("go_value", ""), row.get("difference", ""),
                row.get("source", ""), row.get("confidence", ""),
            ]
            aligns = [_LEFT, _CEN, _CEN, _CEN, _CEN, _LEFT, _CEN]
            for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                cell = ws.cell(r, ci, v)
                cell.font, cell.alignment, cell.border = _NORM, al, _THIN
                cell.fill = sfill if ci == 2 else PatternFill("solid", fgColor="FFFFFF")
            r += 1

        ws.freeze_panes = "A2"
