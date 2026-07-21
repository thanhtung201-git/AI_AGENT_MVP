"""
grid.py — GridSheet: a dense 2-D view of a worksheet with merged cells expanded.

Responsibility: turn a worksheet (or a raw matrix, for tests) into a uniform grid
where merged cells are filled with their top-left value, so downstream detectors
never have to reason about merge ranges or `None` holes from merges.

No parsing logic here — just a clean data structure.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, List, Optional


def col_letter(idx0: int) -> str:
    """0-based column index → Excel letter (0→A, 26→AA)."""
    n = idx0 + 1
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


@dataclass
class GridSheet:
    name: str
    rows: List[List[Any]]   # rows[r][c]; ragged rows are padded on access

    # ── construction ──────────────────────────────────────────────────────────

    @classmethod
    def from_worksheet(cls, ws) -> "GridSheet":
        max_r, max_c = ws.max_row, ws.max_column
        matrix: List[List[Any]] = [
            [effective_text(ws.cell(r, c).value) for c in range(1, max_c + 1)]
            for r in range(1, max_r + 1)
        ]
        # Expand merged ranges: fill every cell of a merge with its top-left value.
        for m in ws.merged_cells.ranges:
            tl = effective_text(ws.cell(m.min_row, m.min_col).value)
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    matrix[r - 1][c - 1] = tl
        return cls(name=ws.title, rows=matrix)

    @classmethod
    def from_matrix(cls, name: str, matrix: List[List[Any]]) -> "GridSheet":
        """For tests: build directly from a raw matrix (already merge-expanded)."""
        return cls(name=name, rows=[list(row) for row in matrix])

    # ── access ────────────────────────────────────────────────────────────────

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def n_cols(self) -> int:
        return max((len(r) for r in self.rows), default=0)

    def cell(self, r: int, c: int) -> Any:
        if 0 <= r < len(self.rows) and 0 <= c < len(self.rows[r]):
            return self.rows[r][c]
        return None

    def row_is_blank(self, r: int) -> bool:
        return all(is_blank(v) for v in self.rows[r])


# ── value helpers (shared) ────────────────────────────────────────────────────

def effective_text(v: Any) -> Any:
    """Resolve a cell value, dropping struck-through runs.

    Buyers revise a cell in place — e.g. a colour cell shows the old code struck
    through and the new code beside it ("G̶1̶ G2"). openpyxl returns the whole
    string "G1 G2"; we keep only the runs that are NOT struck through → "G2".
    Non rich-text values are returned unchanged.
    """
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except Exception:
        return v
    if not isinstance(v, CellRichText):
        return v
    kept: List[str] = []
    for el in v:
        if isinstance(el, TextBlock):
            if getattr(el.font, "strike", None):
                continue  # obsolete value the buyer crossed out
            kept.append(el.text)
        else:
            kept.append(str(el))
    return "".join(kept).strip()


def is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def as_number(v: Any) -> Optional[float]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_number(v: Any) -> bool:
    return as_number(v) is not None


def as_date(v: Any) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return None


def clean_token(v: Any) -> str:
    """Normalize a header/label cell to a short single-line token."""
    if v is None:
        return ""
    s = str(v).strip()
    # size headers often look like "95\n00S" — keep the most size-like line
    parts = [p.strip() for p in s.splitlines() if p.strip()]
    if not parts:
        return ""
    # prefer a part containing letters (size code) over a pure number (e.g. "95")
    for p in parts:
        if any(ch.isalpha() for ch in p):
            return p
    return parts[-1]


def looks_like_article_code(v: Any) -> bool:
    """Generic style/article-code signal: mixes letters and digits, length >= 5.
    Works across brands; short order-type markers ("01", "STOCK") fail it."""
    s = str(v or "").strip()
    if len(s) < 5:
        return False
    return any(ch.isalpha() for ch in s) and any(ch.isdigit() for ch in s)
