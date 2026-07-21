"""
template_writer.py — Fill the company's real Batch GO Upload template in place,
driven by the field rules in config/batch_go_rules.json.

Why this exists
---------------
The company's ERP import reads the Batch GO file against a FIXED schema: exact
sheet names, section labels, column headers and order. A file that merely *looks*
similar (built from scratch) is rejected. So instead of generating a new workbook,
we OPEN the real template, keep every sheet / label / header / format / column
width, DELETE only the sample data rows, and write our values into the right cells
— each cell filled according to its rule.

What the AI agent fills (from the Rules workbook → batch_go_rules.json):
  always        → a fixed constant (Garment Type=W, Trade Term=FOB, …)
  follow_po     → taken from the parsed PO (style, colors, sizes, qty, …)
  seq_go/color/size/lot → a sequential number the code assigns

What the AI agent LEAVES BLANK — the buyer's own ERP fills these on import:
  end_user_fill → filled by their system once the buyer is entered (NOT by us)
  style_desc    → derived from the style description by their system
  no_need_fill  → left blank

Nothing customer-specific lives in this file: sections come from the template,
column positions are found by header name, rules + buyer values live in JSON.
Point BATCH_GO_TEMPLATE / the JSON configs elsewhere and the same code still works.
"""
from __future__ import annotations

import json
import logging
import math
import os
import re
from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.go_compare.canonical import CanonicalOrder

logger = logging.getLogger(__name__)

_CONFIG_DIR   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config")
_RULES_PATH   = os.path.join(_CONFIG_DIR, "batch_go_rules.json")
_BUYER_FILL_PATH = os.path.join(_CONFIG_DIR, "batch_go_buyer_fill.json")

HEADER_SHEET = "Header + main+Color + Size"
LOT_SHEET    = "Lot + BPO"


# ── Rules config (loaded once) ────────────────────────────────────────────────

def _load_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


class RuleBook:
    """Field rules indexed by (sheet, section, field) → rule dict."""

    def __init__(self, rules_path: str = _RULES_PATH, buyer_fill_path: str = _BUYER_FILL_PATH):
        rules = _load_json(rules_path)
        self._by_key: Dict[Tuple[str, str, str], Dict[str, Any]] = {
            (r["sheet"], r["section"], r["field"]): r for r in rules
        }
        # Only these `end_user_fill` fields may take the buyer's fixed value.
        self.buyer_fill = set(_load_json(buyer_fill_path).get("fields") or [])

    def rule(self, sheet: str, section: str, field: str) -> Optional[Dict[str, Any]]:
        return self._by_key.get((sheet, section, field))

    def section_rules(self, sheet: str, section: str) -> List[Dict[str, Any]]:
        return [r for (s, sec, _), r in self._by_key.items() if (s, sec) == (sheet, section)]


# ── Template parsing / editing mechanics ──────────────────────────────────────

@dataclass
class _Section:
    label: str        # section name (":" stripped), e.g. "Header", "Buyer PO"
    header_row: int
    data_last: int


def _parse_sections(ws: Worksheet) -> List[_Section]:
    labels: List[Tuple[int, str]] = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and str(a).strip().endswith(":"):
            labels.append((r, str(a).strip().rstrip(":").strip()))

    sections: List[_Section] = []
    for idx, (lrow, ltext) in enumerate(labels):
        header_row = lrow + 1
        next_boundary = labels[idx + 1][0] if idx + 1 < len(labels) else ws.max_row + 1
        data_last = header_row
        probe = range(1, min(ws.max_column, 6) + 1)
        for r in range(header_row + 1, next_boundary):
            if any(ws.cell(r, c).value not in (None, "") for c in probe):
                data_last = r
        sections.append(_Section(ltext, header_row, data_last))
    return sections


def _header_cols(ws: Worksheet, header_row: int) -> List[Tuple[str, int]]:
    """[(header text, column index)] for a section's header row, in order."""
    out: List[Tuple[str, int]] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v not in (None, ""):
            out.append((str(v).strip(), c))
    return out


def _fill_section(ws: Worksheet, sec: _Section, rows: List[Dict[int, Any]]) -> None:
    """Replace a section's sample data rows with `rows` (col_index → value),
    preserving cell styles. Called bottom-up so edits never shift sections above."""
    first = sec.header_row + 1
    old_n = sec.data_last - sec.header_row
    new_n = len(rows)

    style_tpl: Dict[int, Any] = {}
    if old_n > 0:
        for c in range(1, ws.max_column + 1):
            style_tpl[c] = copy(ws.cell(first, c)._style)

    if new_n < old_n:
        ws.delete_rows(first + new_n, old_n - new_n)
    elif new_n > old_n:
        ws.insert_rows(first + old_n, new_n - old_n)

    for j, rowvals in enumerate(rows):
        rr = first + j
        # Blank the whole row first so reused sample rows never leak old values.
        for c in range(1, ws.max_column + 1):
            ws.cell(rr, c).value = None
            if style_tpl:
                ws.cell(rr, c)._style = copy(style_tpl[c])
        for c, val in rowvals.items():
            if val is not None:
                ws.cell(rr, c).value = val


# ── PO → per-section row contexts (values keyed by exact field name) ───────────

def _distinct(seq):
    seen, out = set(), []
    for x in seq:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _distinct_keys(seq):
    """_distinct for keys that may legitimately be falsy (0, "")."""
    seen, out = set(), []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _apportion(sizes: Dict[str, float]) -> List[Tuple[str, int]]:
    """Round per-size quantities to whole garments while preserving their total.

    A buying sheet can carry fractional per-size quantities (a ratio split) whose
    per-colour total is still whole — e.g. 136.5 + 260 + 169 + 71.5 + 13 = 650.
    Rounding each size on its own makes the sizes stop summing to the lot total,
    and the ERP reconciles Lot Total against the sum of its Buyer PO rows. The
    largest-remainder method hands the leftover units to the biggest fractions, so
    the total stays exact. Ties keep sheet order, so output is reproducible.
    """
    items = [(k, float(v)) for k, v in sizes.items()]
    total = round(sum(v for _, v in items))
    parts = [(k, int(math.floor(v)), v - math.floor(v)) for k, v in items]
    out = {k: f for k, f, _ in parts}
    short = max(total - sum(out.values()), 0)
    for k, _, _frac in sorted(parts, key=lambda x: x[2], reverse=True)[:short]:
        out[k] += 1
    return [(k, out[k]) for k, _ in items]


def _season_year(season: str) -> Optional[str]:
    m = re.search(r"(20\d{2})", season or "")
    if m:
        return m.group(1)
    m = re.match(r"\s*(\d{2})\s*[A-Za-z]{2}", season or "")  # "26FW" → 2026
    if m:
        return "20" + m.group(1)
    return None


def _build_contexts(order: CanonicalOrder) -> Dict[Tuple[str, str], List[Dict[str, Any]]]:
    """One list of field→value contexts per (sheet, section). Only PO-derived and
    sequential values are set here; rules decide constants / buyer defaults later."""
    header_style = str(order.style.value or "").strip()
    season       = order.season.value or ""
    year         = _season_year(season) or ""
    po_number    = order.po_number.value or ""
    order_date0  = order.order_date.value or ""

    def line_style(l):
        return (l.style or header_style or "").strip()

    styles = _distinct(line_style(l) for l in order.lines) or ([header_style] if header_style else [])
    seq_of = {s: i + 1 for i, s in enumerate(styles)}

    header_ctx, main_ctx, color_ctx, size_ctx, lot_ctx, bpo_ctx = ([] for _ in range(6))

    for style in styles:
        seq      = seq_of[style]
        lines_s  = [l for l in order.lines if line_style(l) == style]
        # Order date (발주 / Buyer PO Date) is per style.
        order_date = next((l.order_date for l in lines_s if l.order_date), order_date0)
        # Prod Comp Date is the delivery date. A STOCK run carries no date of its own,
        # so it inherits the dated run of the same style (confirmed by the buyer).
        style_delivery = next((l.delivery_date for l in lines_s if l.delivery_date), "")

        header_ctx.append({
            "GO Seq": seq, "Cust style no": style,
            "Season": season, "Season Year": year,
        })
        main_ctx.append({"GO Seq": seq})

        cseq = 0
        for cc in _distinct((l.color_code or l.color_name) for l in lines_s):
            line = next(l for l in lines_s if (l.color_code or l.color_name) == cc)
            cseq += 1
            color_ctx.append({
                "GO Seq": seq, "SEQ": cseq,
                "Color CODE": line.color_code or "", "Color DESC": line.color_name or "",
                "Cust Color CODE": line.color_code or "", "Cust Color DESC": line.color_name or "",
            })

        sseq = 0
        sizes_s = _distinct(
            sz for l in lines_s for sz in (list(l.size_breakdown.keys()) or ([l.size] if l.size else []))
        )
        for sz in sizes_s:
            sseq += 1
            size_ctx.append({"GO Seq": seq, "Size Code": str(sz).strip(), "SEQ": sseq})

        # ── Lots ──────────────────────────────────────────────────────────────
        # The ERP opens one lot per (run × market): a style ordered twice (a dated
        # run and a STOCK run) gives two runs, and each market inside a run is its
        # own lot. Lot No restarts at 1 for every GO Seq and is the key joining the
        # LOT rows to their Buyer PO rows.
        for lot_no, (blk, dest) in enumerate(
            _distinct_keys((l.block, l.destination) for l in lines_s), start=1
        ):
            lines_l  = [l for l in lines_s if (l.block, l.destination) == (blk, dest)]
            delivery = next((l.delivery_date for l in lines_l if l.delivery_date), "") or style_delivery

            # Build the Buyer PO rows first, then total the lot FROM them, so Lot
            # Total Qty always reconciles against its rows exactly.
            rows = []
            for l in lines_l:
                code  = (l.color_code or l.color_name or "").strip()
                pairs = (_apportion(l.size_breakdown) if l.size_breakdown
                         else ([(l.size, round(l.qty or 0))] if l.size else []))
                for sz, q in pairs:
                    rows.append((code, str(sz).strip(), q))

            lot_ctx.append({
                "GO Seq": seq, "Lot No": lot_no,
                "Total Qty": sum(q for _, _, q in rows) or None,
                "Buyer PO Date": order_date, "Prod Comp Date": delivery,
                "Market": dest,
            })
            for code, sz, q in rows:
                bpo_ctx.append({
                    "GO Seq": seq, "Lot No": lot_no, "Buyer PO No": po_number,
                    "Color Code": code, "Size1": sz, "Qty": q,
                })

    return {
        (HEADER_SHEET, "Header"): header_ctx,
        (HEADER_SHEET, "Main"):   main_ctx,
        (HEADER_SHEET, "Color"):  color_ctx,
        (HEADER_SHEET, "Size1"):  size_ctx,
        (LOT_SHEET,    "LOT"):      lot_ctx,
        (LOT_SHEET,    "Buyer PO"): bpo_ctx,
    }


# ── Rule resolution ───────────────────────────────────────────────────────────

def _resolve(
    rule: Optional[Dict[str, Any]],
    field: str,
    ctx: Dict[str, Any],
    buyer_fill: Optional[set] = None,
) -> Any:
    """Value for one cell given its rule + this row's PO context.

    The AI agent fills: always constants, follow_po (from the PO), the sequential
    numbers, and the handful of `end_user_fill` fields listed in buyer_fill — those
    are fixed per buyer and come from the rules workbook's buyer column. The PO still
    wins when it carries the field itself. Every other `end_user_fill` / `style_desc`
    / `no_need_fill` cell is left blank on purpose.
    """
    if rule is None:
        # Column with no rule (e.g. Size1 'SEQ'): use context value if we have one.
        return ctx.get(field)

    rt = rule["rule"]
    if rt == "always":
        return rule.get("value") or None
    if rt in ("seq_go", "seq_color", "seq_size", "seq_lot"):
        return ctx.get(field)
    if rt == "follow_po":
        v = ctx.get(field)
        return v if v not in (None, "") else None
    if rt == "end_user_fill" and buyer_fill and field in buyer_fill:
        v = ctx.get(field)                       # PO first…
        return v if v not in (None, "") else rule.get("buyer_value")  # …then the buyer
    # end_user_fill, style_desc, no_need_fill → their system fills → leave blank
    return None


# ── Mandatory-field alerts ────────────────────────────────────────────────────

_AGENT_FILLED = {"always", "follow_po", "seq_go", "seq_color", "seq_size", "seq_lot"}


def find_missing_mandatory(
    order: CanonicalOrder, rulebook: Optional[RuleBook] = None
) -> List[Dict[str, str]]:
    """Mandatory fields the agent should fill but the PO gave nothing for.

    Only fields the agent actually fills are reported. A Mandatory cell that the
    buyer's ERP owns (`end_user_fill` outside the buyer-fill list, `no_need_fill`)
    is blank by design and is NOT an alert. Everything reported here means the INPUT
    was short of data, so the end user must supply it before the ERP takes the file.
    """
    rb = rulebook or RuleBook()
    seen: set = set()
    out: List[Dict[str, str]] = []
    for (sheet, section), ctx_list in _build_contexts(order).items():
        for rule in rb.section_rules(sheet, section):
            field = rule["field"]
            if not str(rule.get("mandatory") or "").upper().startswith("Y"):
                continue
            fills = rule["rule"] in _AGENT_FILLED or (
                rule["rule"] == "end_user_fill" and field in rb.buyer_fill
            )
            if not fills:
                continue
            for ctx in ctx_list:
                if _resolve(rule, field, ctx, rb.buyer_fill) not in (None, ""):
                    continue
                key = (sheet, section, field)
                if key not in seen:
                    seen.add(key)
                    out.append({"sheet": sheet, "section": section, "field": field})
                break
    return out


# ── Public API ────────────────────────────────────────────────────────────────

class TemplateBatchGOWriter:
    """Fill the real Batch GO template with a canonical order, per the rule book."""

    def __init__(self, template_path: str, rulebook: Optional[RuleBook] = None):
        self.template_path = template_path
        self.rulebook = rulebook or RuleBook()

    def write(self, order: CanonicalOrder, output_path: str) -> str:
        wb = load_workbook(self.template_path)
        contexts = _build_contexts(order)

        for ws in wb.worksheets:
            sections = _parse_sections(ws)
            for sec in reversed(sections):  # bottom-up keeps upper sections stable
                ctx_list = contexts.get((ws.title, sec.label), [])
                cols     = _header_cols(ws, sec.header_row)
                rows: List[Dict[int, Any]] = []
                for ctx in ctx_list:
                    rowvals: Dict[int, Any] = {}
                    for name, cidx in cols:
                        rule = self.rulebook.rule(ws.title, sec.label, name)
                        val  = _resolve(rule, name, ctx, self.rulebook.buyer_fill)
                        if val is not None:
                            rowvals[cidx] = val
                    rows.append(rowvals)
                _fill_section(ws, sec, rows)

        wb.save(output_path)
        logger.info(f"TemplateBatchGOWriter: filled template → {output_path}")
        return output_path


# ── Structure validation (run before returning the file to the user) ──────────

def _section_signature(ws: Worksheet) -> List[Tuple[str, Tuple[Any, ...]]]:
    sig: List[Tuple[str, Tuple[Any, ...]]] = []
    for sec in _parse_sections(ws):
        headers = tuple(name for name, _ in _header_cols(ws, sec.header_row))
        sig.append((sec.label, headers))
    return sig


def validate_structure(template_path: str, output_path: str) -> List[str]:
    """Return structural mismatches between output and template. Empty = matches."""
    errors: List[str] = []
    tmpl = load_workbook(template_path)
    out  = load_workbook(output_path)

    if tmpl.sheetnames != out.sheetnames:
        errors.append(
            f"Danh sách/thứ tự sheet lệch: template={tmpl.sheetnames} · output={out.sheetnames}"
        )
    for name in tmpl.sheetnames:
        if name not in out.sheetnames:
            continue
        if _section_signature(tmpl[name]) != _section_signature(out[name]):
            errors.append(f"Sheet '{name}': cấu trúc section/cột lệch so với template")
    return errors
