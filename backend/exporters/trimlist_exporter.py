import logging
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Thứ tự nhóm Hazzys
_HAZZYS_CATEGORY_ORDER = [
    "FABRIC/YARN",
    "INTERLINING",
    "THREAD & BUTTON",
    "LABEL",
    "PACKING",
    "OTHER",
]


class TrimlistExporter:
    """Export Trimlist ra file Excel theo định dạng chuẩn ngành may."""

    _HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    _SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
    _GROUP_FILL   = PatternFill("solid", fgColor="BDD7EE")   # xanh nhạt — category header
    _ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
    _TOTAL_FILL   = PatternFill("solid", fgColor="FFF2CC")
    _SOURCE_FILLS = {
        "techpack":    PatternFill("solid", fgColor="E2EFDA"),  # xanh lá nhạt
        "master_trim": PatternFill("solid", fgColor="EBF3FB"),  # xanh nhạt
    }

    _THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    _THICK_BOT = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="medium"),
    )

    # ─────────────────────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────────────────────

    @classmethod
    def export(
        cls,
        trim_items: List[Dict[str, Any]],
        output_path: str,
        meta: Dict[str, Any] = None,
    ) -> str:
        """Export flat (chuẩn cũ) — giữ backward-compat."""
        meta = meta or {}
        wb = Workbook()
        ws = wb.active
        ws.title = "Trimlist"

        row = cls._write_title_block(ws, meta, 1)
        row = cls._write_flat_table(ws, trim_items, row)

        wb.save(output_path)
        logger.info(f"Trimlist exported (flat): {output_path} ({len(trim_items)} items)")
        return output_path

    @classmethod
    def export_hazzys(
        cls,
        trim_items: List[Dict[str, Any]],
        output_path: str,
        meta: Dict[str, Any] = None,
        colors: List[Dict[str, Any]] = None,
    ) -> str:
        """
        Export Hazzys format — grouped by category, color qty columns.

        Args:
            trim_items: từ TrimlistExtractor hoặc MasterTrimReader
            output_path: đường dẫn file .xlsx
            meta:  thông tin PO/style (po_number, style_code, ...)
            colors: list color từ HZSH [{color_code, color_name, total_qty}]
                    dùng để tạo cột qty theo từng màu
        """
        meta   = meta or {}
        colors = colors or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Trimlist"

        row = cls._write_title_block(ws, meta, 1)
        row = cls._write_hazzys_table(ws, trim_items, colors, row)

        wb.save(output_path)
        logger.info(
            f"Trimlist exported (Hazzys): {output_path} "
            f"({len(trim_items)} items, {len(colors)} colors)"
        )
        return output_path

    # ─────────────────────────────────────────────────────────────────────────
    # Internal — title block
    # ─────────────────────────────────────────────────────────────────────────

    @classmethod
    def _write_title_block(cls, ws, meta: Dict[str, Any], start_row: int) -> int:
        row = start_row

        # Tiêu đề
        ws.merge_cells(f"A{row}:K{row}")
        cell = ws[f"A{row}"]
        cell.value     = "TRIM LIST / DANH SÁCH PHỤ LIỆU"
        cell.font      = Font(bold=True, size=14, color="FFFFFF")
        cell.fill      = cls._HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        # Info pairs (3 cột / hàng)
        info_pairs = [
            ("PO Number",   meta.get("po_number", "")),
            ("Style Code",  meta.get("style_code", "")),
            ("Style Name",  meta.get("style_name", "")),
            ("Buyer",       meta.get("buyer", "")),
            ("Season",      meta.get("season", "")),
            ("Order Qty",   meta.get("order_qty", "")),
            ("Factory",     meta.get("factory", "")),
            ("Prepared by", meta.get("prepared_by", "AI Agent")),
            ("Date",        meta.get("date", "")),
        ]
        label_fill = PatternFill("solid", fgColor="D6E4F0")
        for i in range(0, len(info_pairs), 3):
            chunk = info_pairs[i : i + 3]
            col = 1
            for label, value in chunk:
                ws.cell(row, col, label).font = Font(bold=True, size=10)
                ws.cell(row, col, label).fill = label_fill
                ws.cell(row, col + 1, str(value)).font = Font(size=10)
                col += 4
            row += 1

        return row + 1  # khoảng trống

    # ─────────────────────────────────────────────────────────────────────────
    # Internal — flat table (export cũ)
    # ─────────────────────────────────────────────────────────────────────────

    @classmethod
    def _write_flat_table(cls, ws, items: List[Dict], start_row: int) -> int:
        row = start_row
        columns = [
            ("No.",            5),
            ("Trim Item",      22),
            ("Spec / Material",30),
            ("Supplier",       20),
            ("Supplier Code",  18),
            ("Placement",      18),
            ("Qty/Garment",    13),
            ("Unit",            8),
            ("Total Qty",      13),
        ]
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row, col_idx, col_name)
            cell.font      = Font(bold=True, size=10, color="FFFFFF")
            cell.fill      = cls._SUBHDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = cls._THIN
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[row].height = 22
        header_row = row
        row += 1

        for idx, item in enumerate(items, 1):
            fill = cls._ALT_FILL if idx % 2 == 0 else PatternFill()
            values = [
                idx,
                item.get("trim_item", ""),
                item.get("spec", ""),
                item.get("supplier", ""),
                item.get("supplier_code", ""),
                item.get("placement", ""),
                item.get("qty_per_garment", ""),
                item.get("unit", "pc"),
                item.get("total_qty", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row, col_idx, val)
                cell.fill   = fill
                cell.border = cls._THIN
                cell.font   = Font(size=10)
                if col_idx in (1, 7, 8, 9):
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        # Total row
        n_cols = len(columns)
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols - 1)}{row}")
        ws.cell(row, 1, "TOTAL").font      = Font(bold=True, size=10)
        ws.cell(row, 1, "TOTAL").fill      = cls._TOTAL_FILL
        ws.cell(row, 1, "TOTAL").alignment = Alignment(horizontal="right")
        total_qty = sum(
            (it.get("total_qty") or 0) for it in items
            if isinstance(it.get("total_qty"), (int, float))
        )
        ws.cell(row, n_cols, total_qty).font      = Font(bold=True, size=10)
        ws.cell(row, n_cols, total_qty).fill      = cls._TOTAL_FILL
        ws.cell(row, n_cols, total_qty).alignment = Alignment(horizontal="center")
        for c in range(1, n_cols + 1):
            ws.cell(row, c).border = cls._THIN

        ws.freeze_panes = ws.cell(header_row + 1, 1)
        return row + 1

    # ─────────────────────────────────────────────────────────────────────────
    # Internal — Hazzys grouped table
    # ─────────────────────────────────────────────────────────────────────────

    @classmethod
    def _write_hazzys_table(
        cls,
        ws,
        items: List[Dict],
        colors: List[Dict],
        start_row: int,
    ) -> int:
        """
        Viết bảng Hazzys với:
        - Cột cố định: No. | Trim Item | Spec | Supplier | Supplier Code | Placement | Qty/pc | Unit
        - Cột màu: một cột per color (color_code làm header)
        - Cột Total Qty
        - Group header rows theo category (FABRIC/YARN, INTERLINING, ...)
        """
        row = start_row

        # Xây dựng danh sách cột
        fixed_cols = [
            ("No.",            5),
            ("Trim Item",      24),
            ("Spec",           22),
            ("Supplier",       20),
            ("Supplier Code",  16),
            ("Placement",      16),
            ("Qty/pc",         10),
            ("Unit",            8),
        ]
        color_codes = [c.get("color_code", f"C{i+1}") for i, c in enumerate(colors)]
        color_total_qty = {c.get("color_code"): c.get("total_qty", 0) for c in colors}

        all_cols = fixed_cols + [(code, 11) for code in color_codes] + [("Total Qty", 13)]
        n_cols = len(all_cols)
        total_col_idx = n_cols  # 1-based

        # Set column widths
        for col_idx, (_, width) in enumerate(all_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Header row
        for col_idx, (col_name, _) in enumerate(all_cols, 1):
            cell = ws.cell(row, col_idx, col_name)
            cell.font      = Font(bold=True, size=10, color="FFFFFF")
            cell.fill      = cls._SUBHDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = cls._THIN
        ws.row_dimensions[row].height = 24
        header_row = row
        row += 1

        # Group items by category
        grouped: Dict[str, List[Dict]] = {}
        for item in items:
            cat = item.get("category", "OTHER")
            if cat not in grouped:
                grouped[cat] = []
            grouped[cat].append(item)

        item_no = 1
        grand_total = 0.0

        for category in _HAZZYS_CATEGORY_ORDER:
            cat_items = grouped.get(category, [])
            if not cat_items:
                continue

            # Category header row
            ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
            cell = ws.cell(row, 1, category)
            cell.font      = Font(bold=True, size=11, color="1F3864")
            cell.fill      = cls._GROUP_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1)
            cell.border    = cls._THICK_BOT
            ws.row_dimensions[row].height = 20
            row += 1

            cat_total = 0.0

            for item in cat_items:
                qty_per_pc = item.get("qty_per_garment", 1) or 1
                source     = item.get("source", "")

                # Tính qty theo màu nếu có
                color_qtys: List[Optional[float]] = []
                for code in color_codes:
                    cqty = color_total_qty.get(code, 0) or 0
                    color_qtys.append(round(qty_per_pc * cqty, 2) if cqty else None)

                total_item_qty = sum(q for q in color_qtys if q is not None)
                if not total_item_qty:
                    # fallback: dùng total_qty trực tiếp nếu có
                    total_item_qty = item.get("total_qty") or 0
                grand_total += total_item_qty
                cat_total   += total_item_qty

                # Chọn fill theo source
                fill = cls._SOURCE_FILLS.get(source, PatternFill())

                values = [
                    item_no,
                    item.get("trim_item", ""),
                    item.get("spec", item.get("remark", "")),
                    item.get("supplier", ""),
                    item.get("supplier_code", ""),
                    item.get("placement", ""),
                    qty_per_pc,
                    item.get("unit", "pc"),
                ] + (color_qtys or []) + [total_item_qty or ""]

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row, col_idx, val)
                    cell.fill   = fill
                    cell.border = cls._THIN
                    cell.font   = Font(size=10)
                    if col_idx in (1, 7, 8) or col_idx > len(fixed_cols):
                        cell.alignment = Alignment(horizontal="center")

                item_no += 1
                row += 1

            # Category subtotal
            ws.merge_cells(f"A{row}:{get_column_letter(len(fixed_cols))}{row}")
            ws.cell(row, 1, f"Subtotal — {category}").font      = Font(bold=True, size=10, italic=True)
            ws.cell(row, 1).fill      = cls._TOTAL_FILL
            ws.cell(row, 1).alignment = Alignment(horizontal="right")
            ws.cell(row, total_col_idx, round(cat_total, 2)).font      = Font(bold=True, size=10)
            ws.cell(row, total_col_idx).fill      = cls._TOTAL_FILL
            ws.cell(row, total_col_idx).alignment = Alignment(horizontal="center")
            for c in range(1, n_cols + 1):
                ws.cell(row, c).border = cls._THIN
            row += 1

        # Grand total
        ws.merge_cells(f"A{row}:{get_column_letter(total_col_idx - 1)}{row}")
        ws.cell(row, 1, "GRAND TOTAL").font      = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row, 1).fill      = cls._HEADER_FILL
        ws.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, total_col_idx, round(grand_total, 2)).font      = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row, total_col_idx).fill      = cls._HEADER_FILL
        ws.cell(row, total_col_idx).alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, n_cols + 1):
            ws.cell(row, c).border = cls._THIN
        ws.row_dimensions[row].height = 22
        row += 1

        # Legend
        row += 1
        ws.cell(row, 1, "Legend:").font = Font(bold=True, size=9)
        ws.cell(row, 2).fill  = cls._SOURCE_FILLS["techpack"]
        ws.cell(row, 3, "= From Techpack").font = Font(size=9)
        ws.cell(row + 1, 2).fill  = cls._SOURCE_FILLS["master_trim"]
        ws.cell(row + 1, 3, "= From Master Trim").font = Font(size=9)

        ws.freeze_panes = ws.cell(header_row + 1, 1)
        return row + 2
