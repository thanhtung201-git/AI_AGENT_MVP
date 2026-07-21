"""
compare_exporter.py — Export báo cáo so sánh GO vs PO ra Excel.

Sheets:
  1. Summary  — tổng quan (OK / MISMATCH / PARTIAL)
  2. Colors   — chi tiết từng màu (qty GO vs PO, diff, status)
  3. Sizes    — chi tiết size breakdown per màu
"""

import logging
import os
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Colors
_OK_FILL      = PatternFill("solid", fgColor="C6EFCE")   # xanh lá
_WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")   # vàng
_FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")   # đỏ nhạt
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # xanh đậm
_SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")   # xanh nhạt
_SECTION_FILL = PatternFill("solid", fgColor="BDD7EE")

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_STATUS_FILL = {
    "OK":            _OK_FILL,
    "MISMATCH":      _FAIL_FILL,
    "PARTIAL":       _WARN_FILL,
    "MISSING_IN_GO": _FAIL_FILL,
    "MISSING_IN_PO": _WARN_FILL,
}


class CompareExporter:
    """Export báo cáo so sánh GO vs PO."""

    @classmethod
    def export(
        cls,
        compare_result: Dict[str, Any],
        output_path: str = None,
        output_dir: str = None,
    ) -> Dict[str, Any]:
        """
        Tạo file Excel báo cáo so sánh.

        Args:
            compare_result: output của GOCompareService.compare()
            output_path:    đường dẫn file output (ưu tiên)
            output_dir:     thư mục (dùng khi output_path=None)

        Returns:
            {"success": bool, "output_path": str, "error": str|None}
        """
        try:
            if not output_path:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                summary = compare_result.get("summary", {})
                go_no  = summary.get("go_number", "GO")
                po_no  = summary.get("po_number", "PO")
                fname  = f"CompareGOvsPO_{go_no}_{po_no}_{ts}.xlsx"
                out_dir = output_dir or os.path.join("sample_data")
                os.makedirs(out_dir, exist_ok=True)
                output_path = os.path.join(out_dir, fname)

            wb = Workbook()
            cls._write_summary_sheet(wb, compare_result)
            cls._write_colors_sheet(wb, compare_result)
            cls._write_sizes_sheet(wb, compare_result)
            # Xóa sheet mặc định nếu còn
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            wb.save(output_path)
            logger.info(f"CompareExporter: {output_path}")
            return {"success": True, "output_path": output_path, "error": None}

        except Exception as e:
            logger.error(f"CompareExporter error: {e}")
            return {"success": False, "output_path": "", "error": str(e)}

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────

    @classmethod
    def _write_summary_sheet(cls, wb: Workbook, result: Dict[str, Any]) -> None:
        ws = wb.create_sheet("Summary", 0)
        summary     = result.get("summary", {})
        date_check  = result.get("date_check", {})

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 30

        row = 1
        # Title
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value     = "GO vs PO VERIFICATION REPORT"
        cell.font      = Font(bold=True, size=14, color="FFFFFF")
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 2

        status = summary.get("status", "")
        status_fill = _STATUS_FILL.get(status, _WARN_FILL)

        fields = [
            ("Overall Status",     status),
            ("GO Number",          summary.get("go_number", "")),
            ("PO Number",          summary.get("po_number", "")),
            ("Style No",           summary.get("style_no", "")),
            ("GO Total Qty",       summary.get("go_total_qty", 0)),
            ("PO Total Qty",       summary.get("po_total_qty", 0)),
            ("Qty Difference",     summary.get("qty_diff", 0)),
            ("Matched Colors",     summary.get("matched_colors", 0)),
            ("Mismatched Colors",  summary.get("mismatched_colors", 0)),
            ("Missing in GO",      ", ".join(summary.get("missing_in_go", [])) or "—"),
            ("Missing in PO",      ", ".join(summary.get("missing_in_po", [])) or "—"),
            ("GO Ship Date",       date_check.get("go_ship_date", "")),
            ("PO Delivery Date",   date_check.get("po_delivery", "")),
            ("Date Match",         "✓ YES" if date_check.get("date_match") is True
                                   else ("✗ NO" if date_check.get("date_match") is False
                                         else "—")),
        ]

        for label, value in fields:
            lbl_cell = ws.cell(row, 1, label)
            lbl_cell.font   = Font(bold=True, size=10)
            lbl_cell.fill   = _SECTION_FILL
            lbl_cell.border = _THIN

            val_cell = ws.cell(row, 2, value)
            val_cell.font   = Font(size=10)
            val_cell.border = _THIN

            # Highlight status + diff
            if label == "Overall Status":
                val_cell.fill = status_fill
                val_cell.font = Font(bold=True, size=11)
            elif label == "Qty Difference":
                val_cell.fill = _OK_FILL if value == 0 else _FAIL_FILL
            elif label == "Date Match":
                if "YES" in str(value):
                    val_cell.fill = _OK_FILL
                elif "NO" in str(value):
                    val_cell.fill = _FAIL_FILL

            row += 1

        # Generated info
        row += 1
        ws.cell(row, 1, f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=9, italic=True, color="808080")

    # ── Sheet 2: Color Detail ─────────────────────────────────────────────────

    @classmethod
    def _write_colors_sheet(cls, wb: Workbook, result: Dict[str, Any]) -> None:
        ws = wb.create_sheet("Color Detail")
        color_details = result.get("color_details", [])

        headers = [
            ("Color Code",   12),
            ("Color Name",   22),
            ("GO Qty",       12),
            ("PO Qty",       12),
            ("Difference",   13),
            ("Status",       16),
        ]
        for col_idx, (name, width) in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            cell = ws.cell(1, col_idx, name)
            cell.font      = Font(bold=True, size=10, color="FFFFFF")
            cell.fill      = _SUBHDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _THIN
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = ws.cell(2, 1)

        for i, detail in enumerate(color_details, 2):
            status = detail.get("status", "")
            fill   = _STATUS_FILL.get(status, PatternFill())
            diff   = detail.get("qty_diff", 0)

            row_data = [
                detail.get("color_code", ""),
                detail.get("color_name", ""),
                detail.get("go_qty", 0),
                detail.get("po_qty", 0),
                diff,
                status,
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(i, col_idx, val)
                cell.fill   = fill
                cell.border = _THIN
                cell.font   = Font(size=10, bold=(col_idx == 6))
                cell.alignment = Alignment(
                    horizontal="center" if col_idx not in (2,) else "left"
                )

    # ── Sheet 3: Size Detail ──────────────────────────────────────────────────

    @classmethod
    def _write_sizes_sheet(cls, wb: Workbook, result: Dict[str, Any]) -> None:
        ws = wb.create_sheet("Size Detail")
        color_details = result.get("color_details", [])

        # Hàng header gộp
        row = 1
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20

        for detail in color_details:
            size_details = detail.get("size_details", [])
            if not size_details:
                continue

            # Color header
            ws.merge_cells(f"A{row}:F{row}")
            code = detail.get("color_code", "")
            name = detail.get("color_name", "")
            cell = ws.cell(row, 1, f"{code} — {name}")
            cell.font      = Font(bold=True, size=11, color="FFFFFF")
            cell.fill      = _SUBHDR_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1

            # Size table header
            size_headers = ["Size", "GO Qty", "PO Qty", "Difference", "Status"]
            widths = [10, 11, 11, 13, 14]
            for col_idx, (hdr, wid) in enumerate(zip(size_headers, widths), 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    ws.column_dimensions[get_column_letter(col_idx)].width, wid
                )
                cell = ws.cell(row, col_idx, hdr)
                cell.font      = Font(bold=True, size=10)
                cell.fill      = _SECTION_FILL
                cell.border    = _THIN
                cell.alignment = Alignment(horizontal="center")
            row += 1

            # Size rows
            for sd in size_details:
                status = sd.get("status", "")
                fill   = _OK_FILL if status == "OK" else _FAIL_FILL
                for col_idx, val in enumerate([
                    sd.get("size"),
                    sd.get("go_qty", 0),
                    sd.get("po_qty", 0),
                    sd.get("diff", 0),
                    status,
                ], 1):
                    cell = ws.cell(row, col_idx, val)
                    cell.fill   = fill
                    cell.border = _THIN
                    cell.font   = Font(size=10)
                    cell.alignment = Alignment(horizontal="center")
                row += 1

            row += 1  # khoảng cách giữa các màu
