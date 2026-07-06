"""
Export Recap Order Report — Excel màu theo trạng thái OK/WARNING/ERROR.
"""
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Màu sắc
CLR_HEADER_BG  = "1F3864"   # Navy
CLR_HEADER_FG  = "FFFFFF"
CLR_OK         = "E2EFDA"   # Xanh nhạt
CLR_OK_DARK    = "375623"
CLR_WARN       = "FFF2CC"   # Vàng nhạt
CLR_WARN_DARK  = "7D6608"
CLR_ERROR      = "FCE4D6"   # Đỏ nhạt
CLR_ERROR_DARK = "C00000"
CLR_META_BG    = "D6E4F0"
CLR_STRIPE     = "F5F5F5"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


class RecapExporter:

    @staticmethod
    def export(
        checked_items: List[Dict],
        stats: Dict[str, Any],
        meta: Dict[str, Any],
        output_path: str,
    ) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Recap Order"
        ws.sheet_view.showGridLines = False

        row = 1

        # ── TITLE ───────────────────────────────────────────────────────
        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = "RECAP ORDER — ĐỐI CHIẾU ĐƠN ĐẶT HÀNG PHỤ LIỆU"
        ws[f"A{row}"].font      = _font(bold=True, color=CLR_HEADER_FG, size=13)
        ws[f"A{row}"].fill      = _fill(CLR_HEADER_BG)
        ws[f"A{row}"].alignment = _center()
        ws.row_dimensions[row].height = 28
        row += 1

        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   AI Agent MVP"
        ws[f"A{row}"].font      = _font(color=CLR_HEADER_FG, size=9)
        ws[f"A{row}"].fill      = _fill(CLR_HEADER_BG)
        ws[f"A{row}"].alignment = _center()
        ws.row_dimensions[row].height = 16
        row += 1
        row += 1

        # ── META PO ─────────────────────────────────────────────────────
        meta_fields = [
            ("PO Number",   meta.get("po_number",  "")),
            ("Style Code",  meta.get("style_code", "")),
            ("Style Name",  meta.get("style_name", "")),
            ("Buyer",       meta.get("buyer",       "")),
            ("Factory",     meta.get("factory",     "")),
            ("Order Qty",   meta.get("order_qty",   "")),
            ("Season",      meta.get("season",      "")),
            ("Date Ordered",meta.get("date_ordered","")),
        ]
        for i in range(0, len(meta_fields), 2):
            label1, val1 = meta_fields[i]
            label2, val2 = meta_fields[i+1] if i+1 < len(meta_fields) else ("", "")

            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"] = label1
            ws[f"A{row}"].font      = _font(bold=True, size=9)
            ws[f"A{row}"].fill      = _fill(CLR_META_BG)
            ws[f"A{row}"].alignment = _left()

            ws.merge_cells(f"C{row}:E{row}")
            ws[f"C{row}"] = val1
            ws[f"C{row}"].font      = _font(size=9)
            ws[f"C{row}"].alignment = _left()

            ws.merge_cells(f"F{row}:G{row}")
            ws[f"F{row}"] = label2
            ws[f"F{row}"].font      = _font(bold=True, size=9)
            ws[f"F{row}"].fill      = _fill(CLR_META_BG)
            ws[f"F{row}"].alignment = _left()

            ws.merge_cells(f"H{row}:J{row}")
            ws[f"H{row}"] = val2
            ws[f"H{row}"].font      = _font(size=9)
            ws[f"H{row}"].alignment = _left()
            row += 1

        row += 1

        # ── SUMMARY ─────────────────────────────────────────────────────
        summary_bg = "E2EFDA" if stats["passed"] else "FCE4D6"
        verdict    = "✅ ĐẠT — Không có lỗi nghiêm trọng" if stats["passed"] else "❌ KHÔNG ĐẠT — Có lỗi cần kiểm tra"

        ws.merge_cells(f"A{row}:L{row}")
        ws[f"A{row}"] = verdict
        ws[f"A{row}"].font      = _font(bold=True, size=11,
                                        color=CLR_OK_DARK if stats["passed"] else CLR_ERROR_DARK)
        ws[f"A{row}"].fill      = _fill(summary_bg)
        ws[f"A{row}"].alignment = _center()
        ws.row_dimensions[row].height = 22
        row += 1

        # Stats dòng
        stat_labels = [
            ("Tổng items",  stats["total"],    "1F3864", "FFFFFF"),
            ("✅ OK",        stats["ok"],       "375623", "E2EFDA"),
            ("⚠️ Cảnh báo", stats["warnings"], "7D6608", "FFF2CC"),
            ("❌ Lỗi",       stats["errors"],   "C00000", "FCE4D6"),
        ]
        col_map = [("A","B"), ("C","D"), ("E","F"), ("G","H")]
        for (label, val, fg, bg), (c1, c2) in zip(stat_labels, col_map):
            ws.merge_cells(f"{c1}{row}:{c2}{row}")
            ws[f"{c1}{row}"] = f"{label}: {val}"
            ws[f"{c1}{row}"].font      = _font(bold=True, color=fg, size=10)
            ws[f"{c1}{row}"].fill      = _fill(bg)
            ws[f"{c1}{row}"].alignment = _center()
        ws.row_dimensions[row].height = 20
        row += 1
        row += 1

        # ── TABLE HEADER ─────────────────────────────────────────────────
        headers = [
            ("No.",           4),
            ("Supplier Code", 16),
            ("Trim Item",     20),
            ("Spec / Material", 30),
            ("Supplier (Đặt)", 18),
            ("Supplier (Ref)", 18),
            ("Qty Cần",       10),
            ("Qty Đặt",       10),
            ("Unit",          8),
            ("Trạng Thái",    12),
            ("Chi tiết lỗi",  40),
        ]
        for col_idx, (h, _) in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font      = _font(bold=True, color=CLR_HEADER_FG, size=9)
            cell.fill      = _fill(CLR_HEADER_BG)
            cell.alignment = _center()
            cell.border    = _border()
        ws.row_dimensions[row].height = 22
        row += 1

        # ── TABLE DATA ───────────────────────────────────────────────────
        for idx, item in enumerate(checked_items):
            status = item.get("status", "OK")
            bg     = CLR_OK if status == "OK" else (CLR_WARN if status == "WARNING" else CLR_ERROR)
            fg     = CLR_OK_DARK if status == "OK" else (CLR_WARN_DARK if status == "WARNING" else CLR_ERROR_DARK)
            stripe = bg if idx % 2 == 0 else CLR_STRIPE

            qty_req = item.get("qty_required") or ""
            qty_req_str = f"{qty_req:g}" if qty_req else "—"
            qty_ord = item.get("qty_ordered") or 0

            issues_str = " | ".join(item.get("issues") or []) or "—"
            status_str = {"OK": "✅ OK", "WARNING": "⚠️ Cảnh báo", "ERROR": "❌ Lỗi"}.get(status, status)

            row_data = [
                item.get("no", ""),
                item.get("supplier_code", ""),
                item.get("trim_item", ""),
                item.get("spec", ""),
                item.get("supplier", ""),
                item.get("ref_supplier", ""),
                qty_req_str,
                f"{qty_ord:g}" if qty_ord else "0",
                item.get("unit", ""),
                status_str,
                issues_str,
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border    = _border()
                cell.alignment = _center() if col_idx in (1, 7, 8, 9, 10) else _left()
                if status != "OK":
                    cell.fill = _fill(bg)
                    cell.font = _font(size=9, color=fg if col_idx == 10 else "000000",
                                      bold=(col_idx == 10))
                else:
                    cell.fill = _fill(stripe)
                    cell.font = _font(size=9)
            ws.row_dimensions[row].height = 18
            row += 1

        # ── COLUMN WIDTHS ────────────────────────────────────────────────
        for col_idx, (_, width) in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── FREEZE ───────────────────────────────────────────────────────
        # Tìm row header bảng để freeze
        header_row = row - len(checked_items) - 1
        ws.freeze_panes = f"A{header_row + 1}"

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return output_path
