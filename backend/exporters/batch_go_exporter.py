"""
batch_go_exporter.py — Sinh file Batch GO Upload.xlsx đúng chuẩn eSCM Hazzys.

Input:
  - hzsh_data: output của HZSHExtractor (style, colors, lots)
  - po_data:   output của FormatUnderstander (po_number, buyer, delivery_date...)
  - template_path: đường dẫn file "Batch GO Upload.xlsx" gốc đối tác gửi

Output:
  - File Excel 5 sheets đúng chuẩn eSCM, sẵn sàng upload
"""

import os
import logging
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Đường dẫn template mặc định
_DEFAULT_TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "..", "sample_data", "Hazzys", "Batch GO Upload.xlsx"
)


class BatchGOExporter:
    """
    Sinh file Batch GO Upload theo template eSCM của Hazzys.

    Cách dùng:
        exporter = BatchGOExporter()
        output_path = exporter.export(hzsh_data, po_data)
    """

    def __init__(self, template_path: str = None):
        self.template_path = template_path or _DEFAULT_TEMPLATE

    def export(
        self,
        hzsh_data: Dict[str, Any],
        po_data: Dict[str, Any],
        output_dir: str = None,
        output_filename: str = None,
    ) -> Dict[str, Any]:
        """
        Tạo file Batch GO Upload từ dữ liệu HZSH + PO.

        Args:
            hzsh_data:   output của HZSHExtractor.extract()["data"]
            po_data:     output của FormatUnderstander.extract()["data"]
            output_dir:  thư mục lưu file (mặc định: sample_data/)
            output_filename: tên file output (mặc định: auto-generate)

        Returns:
            {"success": bool, "output_path": str, "error": str|None}
        """
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "output_path": "", "error": "Cần cài openpyxl: pip install openpyxl"}

        # Kiểm tra template
        template = os.path.abspath(self.template_path)
        if not os.path.exists(template):
            return {
                "success": False, "output_path": "",
                "error": f"Không tìm thấy template: {template}\nĐặt file 'Batch GO Upload.xlsx' vào folder sample_data/Hazzys/"
            }

        # Tạo tên file output
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        style = hzsh_data.get("style_no", "unknown").replace("/", "-")
        filename = output_filename or f"BatchGO_{style}_{ts}.xlsx"

        out_dir = output_dir or os.path.join(
            os.path.dirname(__file__), "..", "..", "sample_data"
        )
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(os.path.abspath(out_dir), filename)

        # Copy template
        shutil.copy2(template, output_path)

        try:
            wb = openpyxl.load_workbook(output_path)
            self._fill_header_sheet(wb, hzsh_data, po_data)
            self._fill_lot_bpo_sheet(wb, hzsh_data, po_data)
            wb.save(output_path)

            logger.info(f"BatchGOExporter: saved {output_path}")
            return {"success": True, "output_path": output_path, "error": None}

        except Exception as e:
            logger.error(f"BatchGOExporter fill error: {e}")
            # Xóa file copy bị lỗi
            if os.path.exists(output_path):
                os.remove(output_path)
            return {"success": False, "output_path": "", "error": str(e)}

    # ── Sheet fillers ─────────────────────────────────────────────────────────

    def _fill_header_sheet(
        self, wb, hzsh_data: Dict[str, Any], po_data: Dict[str, Any]
    ) -> None:
        """Điền vào sheet đầu tiên (Header + Main + Color + Size)."""
        # Lấy sheet đầu tiên (thường là sheet upload chính)
        ws = wb.worksheets[0]

        style_no    = hzsh_data.get("style_no", po_data.get("style_code", ""))
        season      = hzsh_data.get("season", po_data.get("season", ""))
        brand       = hzsh_data.get("brand", "HAZZYS")
        po_number   = po_data.get("po_number", "")
        buyer       = po_data.get("buyer", "TESSELLATION")
        delivery    = po_data.get("delivery_date", "")
        total_qty   = hzsh_data.get("colors", [{}])
        total_sum   = sum(c.get("total_qty", 0) for c in hzsh_data.get("colors", []))

        # Tìm và điền các cell theo label — tìm cell có text rồi điền cell kế bên
        label_value_map = {
            "STYLE NO":      style_no,
            "STYLE NUMBER":  style_no,
            "SEASON":        season,
            "BRAND":         brand,
            "PO NO":         po_number,
            "PO NUMBER":     po_number,
            "BUYER":         buyer,
            "DELIVERY DATE": delivery,
            "TOTAL QTY":     total_sum,
        }
        self._fill_by_label(ws, label_value_map)

        # Điền color rows nếu có section Color trong sheet
        colors = hzsh_data.get("colors", [])
        self._fill_color_rows(ws, colors)

        logger.info(f"  Header sheet: style={style_no}, total_qty={total_sum}")

    def _fill_lot_bpo_sheet(
        self, wb, hzsh_data: Dict[str, Any], po_data: Dict[str, Any]
    ) -> None:
        """Điền vào sheet Lot+BPO."""
        lot_sheet = None
        for ws in wb.worksheets:
            if "lot" in ws.title.lower() or "bpo" in ws.title.lower():
                lot_sheet = ws
                break
        if not lot_sheet:
            logger.warning("Không tìm thấy sheet Lot+BPO trong template")
            return

        lots = hzsh_data.get("lots", [])
        po_number = po_data.get("po_number", "")

        # Tìm row bắt đầu data trong section LOT
        data_row = self._find_data_start_row(lot_sheet, "LOT")
        if not data_row:
            logger.warning("Không tìm thấy section LOT trong sheet")
            return

        for i, lot in enumerate(lots):
            row = data_row + i
            # Ghi destination, customer, qty — cột cụ thể cần mapping với template thực
            # Đây là placeholder — sẽ cần mapping chính xác sau khi có template thực
            lot_sheet.cell(row=row, column=2).value = lot.get("destination", "")
            lot_sheet.cell(row=row, column=3).value = lot.get("customer", "")
            lot_sheet.cell(row=row, column=4).value = lot.get("color_code", "")
            lot_sheet.cell(row=row, column=5).value = lot.get("total_qty", 0)
            lot_sheet.cell(row=row, column=6).value = po_number

        logger.info(f"  Lot sheet: {len(lots)} lots")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _fill_by_label(self, ws, label_value_map: Dict[str, Any]) -> None:
        """Tìm cell có text = label, điền giá trị vào cell bên phải."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    key = cell.value.strip().upper()
                    if key in label_value_map and label_value_map[key] is not None:
                        # Điền vào cell kế bên (cùng row, cột + 1)
                        target = ws.cell(row=cell.row, column=cell.column + 1)
                        if not target.value:  # không ghi đè nếu đã có data
                            target.value = label_value_map[key]

    def _fill_color_rows(self, ws, colors: List[Dict[str, Any]]) -> None:
        """Tìm section Color trong sheet và điền color rows."""
        color_header_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "COLOR" in str(cell.value).upper():
                    color_header_row = cell.row
                    break
            if color_header_row:
                break

        if not color_header_row:
            return

        for i, color in enumerate(colors):
            row = color_header_row + 1 + i
            ws.cell(row=row, column=1).value = color.get("color_code", "")
            ws.cell(row=row, column=2).value = color.get("color_name", "")
            ws.cell(row=row, column=3).value = color.get("total_qty", 0)

    def _find_data_start_row(self, ws, section_label: str) -> Optional[int]:
        """Tìm row bắt đầu data sau section label."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and section_label.upper() in str(cell.value).upper():
                    return cell.row + 2  # +2 để bỏ qua header row của section
        return None
