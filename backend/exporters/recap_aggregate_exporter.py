import logging
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class RecapAggregateExporter:
    _HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    _SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
    _ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
    _TOTAL_FILL  = PatternFill("solid", fgColor="FFF2CC")
    _THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    @classmethod
    def export(
        cls,
        aggregated_items: List[Dict[str, Any]],
        output_path: str,
        meta: Dict[str, Any] = None,
    ) -> str:
        meta = meta or {}
        wb = Workbook()
        ws = wb.active
        ws.title = "Recap Trim"

        row = 1

        # Tiêu đề
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "TỔNG HỢP PHỤ LIỆU — RECAP TRIM LIST"
        ws[f"A{row}"].font      = Font(bold=True, size=14, color="FFFFFF")
        ws[f"A{row}"].fill      = cls._HEADER_FILL
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1

        # Thông tin tổng hợp
        info = [
            ("Số PO tổng hợp", meta.get("po_count", "")),
            ("Danh sách PO",   meta.get("po_numbers", "")),
            ("Ngày tổng hợp",  meta.get("date", "")),
            ("Người lập",      meta.get("prepared_by", "AI Agent")),
        ]
        for label, val in info:
            if not val:
                continue
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True, size=10)
            ws[f"A{row}"].fill = cls._SUBHDR_FILL
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=10)
            ws.merge_cells(f"B{row}:G{row}")
            ws[f"B{row}"] = str(val)
            ws[f"B{row}"].font = Font(size=10)
            row += 1

        row += 1

        # Header bảng
        headers = ["#", "Trim Item", "Spec / Mô tả", "Nhà cung cấp", "Unit", "Tổng Qty", "Ghi chú (PO)"]
        col_widths = [5, 28, 30, 25, 8, 14, 30]

        for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = cls._SUBHDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = cls._THIN
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 20
        row += 1

        # Dữ liệu
        total_qty_grand = 0
        for idx, item in enumerate(aggregated_items, start=1):
            fill = cls._ALT_FILL if idx % 2 == 0 else PatternFill()
            values = [
                idx,
                item.get("trim_item", ""),
                item.get("spec", ""),
                item.get("supplier", ""),
                item.get("unit", ""),
                item.get("total_qty", 0),
                item.get("po_sources", ""),
            ]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.border    = cls._THIN
                cell.fill      = fill
                cell.font      = Font(size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=(ci in (2, 3, 7)))
                if ci == 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
            total_qty_grand += item.get("total_qty", 0) or 0
            ws.row_dimensions[row].height = 16
            row += 1

        # Total row
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "TỔNG CỘNG"
        ws[f"A{row}"].font      = Font(bold=True, size=10)
        ws[f"A{row}"].fill      = cls._TOTAL_FILL
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row}"].border    = cls._THIN

        ws[f"F{row}"] = total_qty_grand
        ws[f"F{row}"].font          = Font(bold=True, size=10)
        ws[f"F{row}"].fill          = cls._TOTAL_FILL
        ws[f"F{row}"].alignment     = Alignment(horizontal="right", vertical="center")
        ws[f"F{row}"].number_format = "#,##0"
        ws[f"F{row}"].border        = cls._THIN

        ws[f"G{row}"].fill   = cls._TOTAL_FILL
        ws[f"G{row}"].border = cls._THIN
        ws.row_dimensions[row].height = 18

        ws.freeze_panes = f"A{row - len(aggregated_items)}"

        wb.save(output_path)
        logger.info(f"RecapAggregate exported: {output_path} ({len(aggregated_items)} items)")
        return output_path
