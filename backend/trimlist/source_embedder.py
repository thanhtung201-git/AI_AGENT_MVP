"""
source_embedder.py — Make the Trimlist Excel self-contained & click-to-verify.

Goal: the user sends ONLY the trimlist .xlsx; the recipient clicks a "Primary Source"
cell and jumps straight to where the value came from — no external files needed.

Two embedded sources (Cách 1):
  - TrimMaster_src : a verbatim copy of the branch's Trim Master sheet. Coordinates
    are preserved, so a captured cell ref (e.g. "Men Woven!C14") lands exactly.
  - TechPack_src   : the Tech Pack PDF pages that back the values, rendered to PNG
    images and pasted onto a sheet (a PDF has no cells, so we embed the page picture).

Hyperlinks are intra-workbook (=HYPERLINK("#'Sheet'!A1", ...)) so they work offline.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MASTER_SRC_TITLE   = "TrimMaster_src"   # ≤ 31 chars (Excel limit)
TECHPACK_SRC_TITLE = "TechPack_src"


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# ── Master (Excel) — verbatim copy, exact-cell jump ──────────────────────────

def embed_master_sheet(out_wb, master_path: str, sheet_name: str,
                       dest_title: str = MASTER_SRC_TITLE) -> Optional[str]:
    """Copy the branch's Trim Master sheet verbatim into the output workbook so an
    intra-workbook hyperlink can land on the exact source cell. Returns the new
    sheet title, or None if it could not be embedded."""
    if not master_path or not sheet_name or sheet_name == "ALL_SHEETS":
        return None
    try:
        src_wb = load_workbook(master_path, data_only=True, read_only=True)
    except Exception as e:
        logger.warning(f"embed_master_sheet: cannot open master ({e})")
        return None
    try:
        if sheet_name not in src_wb.sheetnames:
            return None
        src = src_wb[sheet_name]
        title = dest_title[:31]
        dest = out_wb.create_sheet(title)
        for row in src.iter_rows():
            for cell in row:
                if cell.value is not None:
                    dest.cell(row=cell.row, column=cell.column, value=cell.value)
        # A light banner so the reader knows what this sheet is.
        dest.sheet_properties.tabColor = "9BBB59"
        logger.info(f"embed_master_sheet: copied '{sheet_name}' → '{title}'")
        return title
    except Exception as e:
        logger.warning(f"embed_master_sheet failed: {e}")
        return None
    finally:
        src_wb.close()


# ── Tech Pack (PDF) — page images, jump to the page picture ──────────────────

def embed_techpack_pages(out_wb, pdf_path: str, items: List[Any],
                         dest_title: str = TECHPACK_SRC_TITLE) -> Dict[str, str]:
    """Find which PDF page backs each item, render those pages to images, paste them
    onto one sheet, and return {normalized_key: anchor_cell} so the traceability writer
    can hyperlink each row to its page picture.

    `items` is a list of (material_name, material_code). We locate a page by CODE first
    (unique → precise page) and fall back to the name. Anchors are stored under both
    keys. Only PDF Tech Packs are handled (a picture makes no sense for xlsx/docx)."""
    if not pdf_path or not pdf_path.lower().endswith(".pdf"):
        return {}
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        logger.warning(f"embed_techpack_pages: PyMuPDF unavailable ({e})")
        return {}

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logger.warning(f"embed_techpack_pages: cannot open PDF ({e})")
        return {}

    try:
        page_norm = [_norm(doc[i].get_text()) for i in range(len(doc))]

        def _find(term: str) -> Optional[int]:
            nt = _norm(term)
            if not nt or len(nt) < 5:
                return None
            for pi, ptext in enumerate(page_norm):
                if nt in ptext:
                    return pi
            return None

        # key → page. Prefer the code (unique) over the name (may be generic).
        key_page: Dict[str, int] = {}
        needed_pages: set = set()
        for it in items:
            name, code = (it if isinstance(it, (list, tuple)) else (it, None))
            page = _find(code) if code else None
            if page is None:
                page = _find(name)
            if page is None:
                continue
            needed_pages.add(page)
            for k in (code, name):
                if k and _norm(k):
                    key_page[_norm(k)] = page

        if not needed_pages:
            return {}

        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage

        ws = out_wb.create_sheet(dest_title[:31])
        ws.sheet_properties.tabColor = "4F81BD"

        page_anchor: Dict[int, str] = {}
        row_cursor = 1
        for pi in sorted(needed_pages):
            anchor = f"A{row_cursor}"
            ws.cell(row=row_cursor, column=1, value=f"=== Tech Pack — Trang {pi + 1} ===")
            page_anchor[pi] = anchor
            pix = doc[pi].get_pixmap(matrix=fitz.Matrix(1.6, 1.6))
            pil = PILImage.open(io.BytesIO(pix.tobytes("png")))
            img = XLImage(pil)
            img.anchor = f"A{row_cursor + 1}"
            ws.add_image(img)
            # leave room below the image (approx by pixel height / row height)
            row_cursor += 2 + int(pil.height / 18) + 3

        # Map each key to its page's anchor
        key_anchor = {k: page_anchor[pi] for k, pi in key_page.items() if pi in page_anchor}
        logger.info(f"embed_techpack_pages: embedded {len(needed_pages)} page(s), "
                    f"{len(key_anchor)} keys mapped")
        return key_anchor
    except Exception as e:
        logger.warning(f"embed_techpack_pages failed: {e}")
        return {}
    finally:
        doc.close()


def value_key(*parts: Any) -> str:
    """Normalized key used to look a row up in the techpack value→anchor map."""
    return _norm(" ".join(str(p) for p in parts if p))
