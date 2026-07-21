"""
excel_writer.py — Write Trimlist to Excel matching the standard HAZZYS format.

Sheet layout (matches reference file Trimlist HZSH6C331 S26M01565.xlsx):
  Row 1       : Style + PO header (full-width merged)
  Row 2-3     : Column headers with BODY COLOR sub-headers per colorway
  Category row: Full-width merged, category name
  Data rows   : A:B=Name, C=Code, D:E=Supplier, F=QTY, G:H=Placement,
                I=Composition, J[/K/…]=Color per colorway, Last=Remark
Sheet 2 — TRACEABILITY
Sheet 3 — ALERTS + Self-check
"""
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.trimlist.traceability import TrimRow, CATEGORY_ORDER
from backend.trimlist.validator import Alert

logger = logging.getLogger(__name__)

# ── Palette (matching sample file aesthetic) ──────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy — title row
_SUBHDR    = PatternFill("solid", fgColor="2E75B6")   # medium blue — col headers
_CAT_FILLS = {
    "FABRIC/YARN":     PatternFill("solid", fgColor="D9E1F2"),
    "INTERLINING":     PatternFill("solid", fgColor="E2EFDA"),
    "THREAD & BUTTON": PatternFill("solid", fgColor="FFF2CC"),
    "LABEL":           PatternFill("solid", fgColor="FCE4D6"),
    "PACKING":         PatternFill("solid", fgColor="F4CCFF"),
    "OTHER":           PatternFill("solid", fgColor="F2F2F2"),
}
_ALT_FILL  = PatternFill("solid", fgColor="F9FBFF")
_TBD_FILL  = PatternFill("solid", fgColor="FFE0E0")
_ALERT_FILLS = {
    "ERROR":   PatternFill("solid", fgColor="FFD7D7"),
    "WARNING": PatternFill("solid", fgColor="FFF3CC"),
    "INFO":    PatternFill("solid", fgColor="EEF4FF"),
}

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_THICK_BOT = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="medium"),
)

_WHITE  = Font(color="FFFFFF", bold=True, size=10)
_BOLD   = Font(bold=True, size=10)
_NORM   = Font(size=9)
_SMALL  = Font(size=8, color="555555")
_TBD_FN = Font(size=9, color="CC0000", italic=True)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _c(ws, row: int, col: int, value: Any = "", **style):
    cell = ws.cell(row=row, column=col, value=value)
    if "fill"   in style: cell.fill      = style["fill"]
    if "font"   in style: cell.font      = style["font"]
    if "border" in style: cell.border    = style["border"]
    if "align"  in style: cell.alignment = style["align"]
    return cell


class TrimlistExcelWriter:
    """Writes the final Trimlist to an Excel workbook matching the sample format."""

    def write(
        self,
        rows: List[TrimRow],
        alerts: List[Alert],
        output_path: str,
        meta: Optional[Dict] = None,
        email_changes: Optional[List[str]] = None,
        self_check: Optional[List[str]] = None,
        master_path: Optional[str] = None,
        master_sheet: Optional[str] = None,
        techpack_path: Optional[str] = None,
    ) -> str:
        meta          = meta or {}
        email_changes = email_changes or []
        self_check    = self_check or []

        wb = Workbook()

        # Collect all colorways present in data
        all_colorways = sorted({k for row in rows for k in (row.colors or {}).keys()})

        self._write_trimlist_sheet(wb, rows, meta, all_colorways)

        # Embed the sources so the file is self-contained and click-to-verify.
        from backend.trimlist.source_embedder import embed_master_sheet, embed_techpack_pages
        master_src = embed_master_sheet(wb, master_path, master_sheet) if master_path else None
        tp_anchor  = embed_techpack_pages(
            wb, techpack_path, [(r.material_name, r.material_code) for r in rows]
        ) if techpack_path else {}

        self._write_traceability_sheet(wb, rows, master_src, tp_anchor)
        self._write_alerts_sheet(wb, alerts, email_changes, self_check)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        logger.info(f"TrimlistExcelWriter: saved to {output_path}")
        return output_path

    # ── Sheet 1: Trimlist ─────────────────────────────────────────────────────

    def _write_trimlist_sheet(
        self, wb: Workbook, rows: List[TrimRow], meta: Dict, all_colorways: list
    ):
        ws = wb.create_sheet("Trimlist", 0)
        ws.sheet_view.showGridLines = False

        # ── Column layout (matches sample) ─────────────────────────────────
        # Fixed cols: A:B=Name, C=Code, D:E=Supplier, F=QTY, G:H=Placement,
        #             I=Composition, then 1 col per colorway, last col=Remark
        # Column indices (1-based):
        COL_NAME    = 1   # A — also spans B
        COL_CODE    = 3   # C
        COL_SUPP    = 4   # D — also spans E
        COL_QTY     = 6   # F
        COL_PLACE   = 7   # G — also spans H
        COL_COMP    = 9   # I
        COL_COLOR0  = 10  # J — first colorway (or single "BODY COLOR")
        # subsequent colorways: 11, 12, ...
        n_cw        = max(len(all_colorways), 1)
        COL_REMARK  = COL_COLOR0 + n_cw   # after all colorway cols

        total_cols  = COL_REMARK
        last_col_ltr = get_column_letter(total_cols)

        # Column widths
        ws.column_dimensions["A"].width = 28   # Name part 1
        ws.column_dimensions["B"].width = 4    # Name part 2 (merged with A)
        ws.column_dimensions["C"].width = 18   # Code
        ws.column_dimensions["D"].width = 18   # Supplier part 1
        ws.column_dimensions["E"].width = 4    # Supplier part 2 (merged)
        ws.column_dimensions["F"].width = 10   # QTY
        ws.column_dimensions["G"].width = 22   # Placement part 1
        ws.column_dimensions["H"].width = 4    # Placement part 2 (merged)
        ws.column_dimensions["I"].width = 20   # Composition
        for i in range(n_cw):
            ws.column_dimensions[get_column_letter(COL_COLOR0 + i)].width = 14
        ws.column_dimensions[last_col_ltr].width = 30  # Remark

        # ── Row 1: Style / PO title ─────────────────────────────────────────
        r = 1
        ws.row_dimensions[r].height = 20
        style_code = meta.get("style_code") or ""
        po_num     = meta.get("po_number") or ""
        title      = f"{style_code}  {po_num}".strip() or "TRIM LIST / BILL OF MATERIALS"
        ws.merge_cells(f"A{r}:{last_col_ltr}{r}")
        _c(ws, r, 1, title,
           fill=_HDR_FILL, font=Font(color="FFFFFF", bold=True, size=12),
           align=_CENTER, border=_THIN)

        # ── Row 2-3: Column headers ─────────────────────────────────────────
        r = 2
        ws.row_dimensions[r].height = 22
        ws.row_dimensions[3].height = 18

        def _hdr2(col: int, label: str, span_to: int = None, subrow_label: str = None):
            """Write a 2-row header cell (merged rows 2-3) or 2-level header."""
            end_col = span_to or col
            end_ltr = get_column_letter(end_col)
            col_ltr = get_column_letter(col)
            if subrow_label is None:
                # Single-level: merge rows 2-3
                ws.merge_cells(f"{col_ltr}2:{end_ltr}3")
                _c(ws, 2, col, label,
                   fill=_SUBHDR, font=_WHITE, align=_CENTER, border=_THIN)
            else:
                # Two-level: row 2 = group label, row 3 = sub-label
                _c(ws, 2, col, label,
                   fill=_SUBHDR, font=_WHITE, align=_CENTER, border=_THIN)
                _c(ws, 3, col, subrow_label,
                   fill=_SUBHDR, font=Font(color="FFFFFF", size=9), align=_CENTER, border=_THIN)

        # Fixed headers (rows 2-3 merged)
        _hdr2(COL_NAME,  "MATERIAL NAME", span_to=COL_NAME + 1)
        _hdr2(COL_CODE,  "CODE")
        _hdr2(COL_SUPP,  "SUPPLIER",      span_to=COL_SUPP + 1)
        _hdr2(COL_QTY,   "Q'TY")
        _hdr2(COL_PLACE, "PLACEMENT",     span_to=COL_PLACE + 1)
        _hdr2(COL_COMP,  "COMPOSITION")
        _hdr2(COL_REMARK,"REMARK")

        # Colorway headers
        if n_cw > 1:
            # "BODY COLOR" merged across all colorway columns in row 2
            cw_end = get_column_letter(COL_COLOR0 + n_cw - 1)
            ws.merge_cells(f"{get_column_letter(COL_COLOR0)}2:{cw_end}2")
            _c(ws, 2, COL_COLOR0, "BODY COLOR",
               fill=_SUBHDR, font=_WHITE, align=_CENTER, border=_THIN)
            # Individual colorway names in row 3
            for i, cw in enumerate(all_colorways):
                _c(ws, 3, COL_COLOR0 + i, cw,
                   fill=_SUBHDR, font=Font(color="FFFFFF", size=9),
                   align=_CENTER, border=_THIN)
        else:
            _hdr2(COL_COLOR0, "BODY COLOR")

        # ── Data rows grouped by category ───────────────────────────────────
        r = 4
        current_cat = None

        def _merge_name_row(row_r: int, value: str, fill=None, font=None, is_cat=False):
            """Write A:B merged cell for name/category."""
            ws.merge_cells(f"A{row_r}:B{row_r}")
            _c(ws, row_r, COL_NAME, value,
               fill=fill or _ALT_FILL,
               font=font or _NORM,
               align=_LEFT, border=_THIN if not is_cat else _THICK_BOT)

        for row in rows:
            # Category separator
            if row.category != current_cat:
                current_cat = row.category
                ws.row_dimensions[r].height = 18
                cat_fill = _CAT_FILLS.get(current_cat, _CAT_FILLS["OTHER"])
                ws.merge_cells(f"A{r}:{last_col_ltr}{r}")
                _c(ws, r, 1, current_cat,
                   fill=cat_fill,
                   font=Font(bold=True, size=10),
                   align=_LEFT, border=_THICK_BOT)
                r += 1

            ws.row_dimensions[r].height = 32
            cat_fill = _CAT_FILLS.get(row.category, _ALT_FILL)

            # A:B — Material name (merged)
            ws.merge_cells(f"A{r}:B{r}")
            _c(ws, r, COL_NAME, row.material_name or "",
               fill=cat_fill, font=_NORM, align=_LEFT, border=_THIN)

            # C — Code
            code = row.material_code or ""
            code_fill = _TBD_FILL if not code else cat_fill
            code_font = _TBD_FN   if not code else _NORM
            _c(ws, r, COL_CODE, code or "TBD",
               fill=code_fill, font=code_font, align=_CENTER, border=_THIN)

            # D:E — Supplier (merged)
            ws.merge_cells(f"D{r}:E{r}")
            supp = row.supplier or ""
            supp_fill = _TBD_FILL if not supp else cat_fill
            supp_font = _TBD_FN   if not supp else _NORM
            _c(ws, r, COL_SUPP, supp or "TBD",
               fill=supp_fill, font=supp_font, align=_LEFT, border=_THIN)

            # F — QTY (consumption, e.g. "1ea", "11+1")
            qty = row.consumption or ""
            _c(ws, r, COL_QTY, qty,
               fill=cat_fill, font=_NORM, align=_CENTER, border=_THIN)

            # G:H — Placement (merged)
            ws.merge_cells(f"G{r}:H{r}")
            _c(ws, r, COL_PLACE, row.placement or "",
               fill=cat_fill, font=_NORM, align=_LEFT, border=_THIN)

            # I — Composition / spec
            _c(ws, r, COL_COMP, row.spec or "",
               fill=cat_fill, font=_NORM, align=_LEFT, border=_THIN)

            # Colorway color columns
            if n_cw > 1 and row.colors:
                for i, cw in enumerate(all_colorways):
                    color_val = row.colors.get(cw, "")
                    _c(ws, r, COL_COLOR0 + i, color_val,
                       fill=cat_fill, font=_NORM, align=_CENTER, border=_THIN)
            else:
                # Single colorway or no colorways
                color_val = row.color or ""
                if not color_val and row.colors:
                    color_val = list(row.colors.values())[0] if row.colors else ""
                _c(ws, r, COL_COLOR0, color_val,
                   fill=cat_fill, font=_NORM, align=_CENTER, border=_THIN)

            # Remark
            _c(ws, r, COL_REMARK, row.remark or "",
               fill=cat_fill, font=_SMALL, align=_LEFT, border=_THIN)

            # Alert indicator: tint the name cell
            if row.alerts:
                has_err = any("ERROR" in a for a in row.alerts)
                tint    = PatternFill("solid", fgColor="FFD7D7" if has_err else "FFF3CC")
                ws.cell(row=r, column=COL_NAME).fill = tint

            r += 1

        ws.freeze_panes = "A4"

    # ── Sheet 2: Traceability ─────────────────────────────────────────────────

    def _write_traceability_sheet(self, wb: Workbook, rows: List[TrimRow],
                                  master_src: Optional[str] = None,
                                  tp_anchor: Optional[Dict[str, str]] = None):
        from backend.trimlist.source_embedder import value_key, TECHPACK_SRC_TITLE
        tp_anchor = tp_anchor or {}
        ws = wb.create_sheet("TRACEABILITY", 1)
        ws.sheet_view.showGridLines = False

        headers = ["#", "Material Name", "Category", "Tech Pack Ref",
                   "Trim Master Ref", "Buyer Rule", "Email Override", "Primary Source"]
        widths  = [4, 28, 18, 32, 32, 28, 28, 16]

        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 28
        for ci, h in enumerate(headers, 1):
            _c(ws, 1, ci, h, fill=_HDR_FILL, font=_WHITE, align=_CENTER, border=_THIN)

        _src_fill = {
            "EMAIL":       PatternFill("solid", fgColor="FFEEDD"),
            "BUYER_RULE":  PatternFill("solid", fgColor="E8F5E8"),
            "TRIM_MASTER": PatternFill("solid", fgColor="E8F0FF"),
            "TECH_PACK":   PatternFill("solid", fgColor="F9FBFF"),
            "UNKNOWN":     PatternFill("solid", fgColor="F5F5F5"),
        }
        _link_font = Font(size=9, color="0563C1", underline="single")

        for i, row in enumerate(rows, 2):
            ws.row_dimensions[i].height = 22
            src     = row.source
            primary = src.primary_source()
            values  = [i - 1, row.material_name, row.category,
                       src.techpack_ref or "", src.master_ref or "",
                       src.buyer_rule or "", src.email_ref or ""]
            aligns  = [_CENTER, _LEFT, _CENTER, _LEFT, _LEFT, _LEFT, _LEFT]
            rfill   = _src_fill.get(primary, _ALT_FILL)
            for ci, (val, aln) in enumerate(zip(values, aligns), 1):
                _c(ws, i, ci, val, fill=rfill, font=_SMALL, align=aln, border=_THIN)

            # Primary Source (col 8) → intra-workbook hyperlink to the embedded source
            tp_keys = [value_key(row.material_code), value_key(row.material_name)]
            target  = self._source_link(src, primary, master_src, tp_anchor,
                                        tp_keys, TECHPACK_SRC_TITLE)
            pcell = ws.cell(row=i, column=8)
            pcell.fill, pcell.alignment, pcell.border = rfill, _CENTER, _THIN
            if target:
                pcell.value = f'=HYPERLINK("{target}","{primary} ▸")'
                pcell.font  = _link_font
            else:
                pcell.value = primary
                pcell.font  = _SMALL

        ws.freeze_panes = "A2"

    @staticmethod
    def _source_link(src, primary: str, master_src: Optional[str],
                     tp_anchor: Dict[str, str], tp_keys: List[str], tp_title: str) -> Optional[str]:
        """Build the intra-workbook hyperlink target for a row's primary source."""
        if primary == "TRIM_MASTER" and master_src and getattr(src, "master_loc", None):
            cell = (src.master_loc or {}).get("cell")
            if cell:
                return f"#'{master_src}'!{cell}"
        if primary == "TECH_PACK":
            for k in tp_keys:
                if k and tp_anchor.get(k):
                    return f"#'{tp_title}'!{tp_anchor[k]}"
        return None

    # ── Sheet 3: Alerts ───────────────────────────────────────────────────────

    def _write_alerts_sheet(
        self,
        wb: Workbook,
        alerts: List[Alert],
        email_changes: List[str],
        self_check: List[str],
    ):
        ws = wb.create_sheet("ALERTS", 2)
        ws.sheet_view.showGridLines = False

        headers = ["#", "Severity", "Item Name", "Alert Code", "Message"]
        widths  = [4, 12, 25, 22, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary row (inserted first, populated at end)
        ws.row_dimensions[1].height = 20

        r = 2
        ws.row_dimensions[r].height = 28
        for ci, h in enumerate(headers, 1):
            _c(ws, r, ci, h, fill=_HDR_FILL, font=_WHITE, align=_CENTER, border=_THIN)

        r = 3
        for i, alert in enumerate(alerts, 1):
            ws.row_dimensions[r].height = 22
            fill   = _ALERT_FILLS.get(alert.severity, _ALT_FILL)
            values = [i, alert.severity, alert.item_name, alert.code, alert.message]
            aligns = [_CENTER, _CENTER, _LEFT, _CENTER, _LEFT]
            for ci, (val, aln) in enumerate(zip(values, aligns), 1):
                _c(ws, r, ci, val,
                   fill=fill,
                   font=Font(bold=(alert.severity == "ERROR"), size=9),
                   align=aln, border=_THIN)
            r += 1

        # Email changes section
        if email_changes:
            r += 1
            ws.merge_cells(f"A{r}:E{r}")
            _c(ws, r, 1, "EMAIL / NOTE CHANGES APPLIED",
               fill=PatternFill("solid", fgColor="FFF3CC"),
               font=Font(bold=True, size=10), align=_LEFT, border=_THIN)
            r += 1
            for change in email_changes:
                ws.merge_cells(f"B{r}:E{r}")
                _c(ws, r, 1, "•", font=_NORM, align=_CENTER, border=_THIN)
                _c(ws, r, 2, change, font=_SMALL, align=_LEFT, border=_THIN)
                r += 1

        # Self-check section
        if self_check:
            r += 1
            ws.merge_cells(f"A{r}:E{r}")
            _c(ws, r, 1, "SELF-CHECK REPORT",
               fill=_HDR_FILL,
               font=Font(bold=True, size=10, color="FFFFFF"),
               align=_LEFT, border=_THIN)
            r += 1
            _icon_fill = {"✓": "E8F5E8", "⚠": "FFF3CC", "ℹ": "EEF4FF"}
            for item in self_check:
                icon  = item[0] if item else ""
                fclr  = _icon_fill.get(icon, "F9FBFF")
                pfill = PatternFill("solid", fgColor=fclr)
                ws.merge_cells(f"B{r}:E{r}")
                _c(ws, r, 1, icon,  fill=pfill, font=Font(bold=True, size=10), align=_CENTER, border=_THIN)
                _c(ws, r, 2, item[2:].strip() if len(item) > 2 else item,
                   fill=pfill, font=_SMALL, align=_LEFT, border=_THIN)
                r += 1

        # Summary row 1
        errors   = sum(1 for a in alerts if a.severity == "ERROR")
        warnings = sum(1 for a in alerts if a.severity == "WARNING")
        summary  = (f"VALIDATION SUMMARY — {errors} Error(s) · {warnings} Warning(s) "
                    f"· {len(email_changes)} Email change(s)")
        ws.merge_cells("A1:E1")
        _c(ws, 1, 1, summary,
           fill=PatternFill("solid", fgColor="FFD7D7" if errors else "FFF3CC" if warnings else "E8F5E8"),
           font=Font(bold=True, size=10),
           align=_CENTER, border=_THIN)

        ws.freeze_panes = "A3"
