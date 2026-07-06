"""
POST /api/agent/run — Upload file PO → AI trích xuất → tự tìm techpack → tạo trimlist.
Đây là luồng thống nhất, tương đương run_agent.py.
"""
import os
import glob
import logging
import asyncio
import traceback
from datetime import datetime

from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel

from backend.agents.po_agent import POAgent
from backend.exporters.json_exporter import JsonExporter
from backend.exporters.excel_exporter import ExcelExporter
from backend.normalization.mapper import DataMapper
from backend.schemas.canonical import CanonicalSchema
from backend.database.supabase_client import SupabaseClient
from backend.extractors.trimlist_extractor import TrimlistExtractor
from backend.exporters.trimlist_exporter import TrimlistExporter
from backend.services.scan_service import (
    get_new_files, get_processed_log, get_all_po_files,
    mark_processed, mark_failed, reset_file, SCAN_FOLDER,
)
from tools.reader import read_file

logger = logging.getLogger(__name__)
router = APIRouter()

UPLOAD_DIR  = "sample_data/uploads"
OUTPUT_DIR  = "sample_data"
TECHPACK_DIR = "Teck_pack"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _find_techpack(style_code: str) -> list:
    all_files = (
        glob.glob(f"{TECHPACK_DIR}/*.pdf") +
        glob.glob(f"{TECHPACK_DIR}/*.xlsx") +
        glob.glob(f"{TECHPACK_DIR}/*.docx")
    )
    if not all_files:
        return []
    if not style_code:
        return all_files

    style_lower = style_code.lower()
    by_name = [f for f in all_files if style_lower in os.path.basename(f).lower()]
    if by_name:
        return by_name

    by_content = []
    for f in all_files:
        try:
            r = read_file(f)
            if r["success"] and style_lower in r["text"].lower():
                by_content.append(f)
        except Exception:
            pass
    return by_content or all_files


@router.post("/run")
async def run_agent(file: UploadFile = File(...)):
    """
    Upload file PO → trích xuất PO → tự tìm techpack khớp → tạo trimlist.
    Trả về kết quả PO + trimlist + đường dẫn download.
    """
    ext = os.path.splitext(file.filename)[-1].lower()
    ALLOWED = {".pdf", ".xlsx", ".xls", ".xlsm", ".docx", ".doc",
               ".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp",
               ".eml", ".msg"}
    if ext not in ALLOWED:
        raise HTTPException(400, f"Định dạng '{ext}' không hỗ trợ. Chấp nhận: PDF, Excel, Word, Ảnh, Email")

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    saved_path = os.path.join(UPLOAD_DIR, f"po_{timestamp}{ext}")
    contents   = await file.read()
    with open(saved_path, "wb") as f:
        f.write(contents)

    # ── Bước 1-2: PO Agent ──────────────────────────────────────────────────
    try:
        agent  = POAgent()
        result = agent.process_request(
            user_request="Extract PO từ file và lưu vào database",
            file_path=saved_path,
        )
    except Exception as e:
        raise HTTPException(500, f"PO Agent lỗi: {e}")

    if result["status"] != "success":
        raise HTTPException(422, f"Trích xuất PO thất bại: {result.get('reason', 'Unknown')}")

    results = result.get("results", {})
    header  = results.get("header_extractor", {}).get("header", {})
    items   = results.get("item_extractor",  {}).get("items",  [])

    combined   = {**header, "items": items,
                  "total_quantity_all": sum(i.get("total_quantity") or 0 for i in items),
                  "total_amount":       sum(i.get("total_price")    or 0.0 for i in items)}
    normalized = DataMapper.map_po_data(combined)
    po_model   = CanonicalSchema.validate_and_load(normalized)

    # Export PO files
    json_path  = os.path.join(OUTPUT_DIR, f"output_po_{timestamp}.json")
    excel_path = os.path.join(OUTPUT_DIR, f"output_po_{timestamp}.xlsx")
    JsonExporter.export(po_model, json_path)
    ExcelExporter.export(po_model, excel_path)

    # Lưu Supabase
    po_id = None
    try:
        db    = SupabaseClient()
        po_id = db.insert_po(header=header, items=items)
    except Exception as e:
        logger.warning(f"Supabase bỏ qua: {e}")

    style_code  = po_model.items[0].style_code if po_model.items else ""
    order_qty   = po_model.total_quantity_all or sum(i.total_quantity or 0 for i in po_model.items)
    _item0      = po_model.items[0] if po_model.items else None
    style_name  = getattr(po_model, "style_name", "") or (getattr(_item0, "style_name", "") if _item0 else "")
    factory     = getattr(po_model, "factory", "") or ""

    po_data = {
        "po_id":        po_id,
        "po_number":    po_model.po_number,
        "style_code":   style_code,
        "style_name":   style_name,
        "buyer":        po_model.buyer,
        "factory":      factory,
        "total_qty":    order_qty,
        "total_amount": po_model.total_amount,
        "item_count":   len(po_model.items),
        "items":        [i.model_dump() for i in po_model.items],
        "excel_path":   excel_path,
    }

    # ── Bước 3: Tìm techpack & tạo trimlist ────────────────────────────────
    techpack_files = _find_techpack(style_code)
    if not techpack_files:
        return {
            "status":           "partial",
            "timestamp":        timestamp,
            "po":               po_data,
            "trimlist":         None,
            "techpack_found":   [],
            "warning":          f"Không tìm thấy file techpack trong thư mục '{TECHPACK_DIR}/'",
        }

    meta = {
        "po_number":   po_model.po_number  or "",
        "style_code":  style_code,
        "style_name":  style_name,
        "buyer":       po_model.buyer      or "",
        "order_qty":   f"{order_qty:,} pcs",
        "factory":     factory,
        "season":      getattr(po_model, "season", "") or "",
        "date":        datetime.now().strftime("%d/%m/%Y"),
        "prepared_by": "AI Agent",
    }

    extractor      = TrimlistExtractor()
    all_trim_items = []
    techpack_names = []

    for tp_path in techpack_files:
        techpack_names.append(os.path.basename(tp_path))
        r = read_file(tp_path)
        if not r.get("success"):
            logger.warning(f"Đọc techpack thất bại: {tp_path}")
            continue
        try:
            items_trim = extractor.extract(r["text"], order_qty=order_qty)
            for it in items_trim:
                it["_source_file"] = os.path.basename(tp_path)
            all_trim_items.extend(items_trim)
        except RuntimeError as e:
            if "RATE LIMIT" in str(e):
                raise HTTPException(429, str(e))
            logger.error(f"Trim extract lỗi: {e}")

    if not all_trim_items:
        return {
            "status":         "partial",
            "timestamp":      timestamp,
            "po":             po_data,
            "trimlist":       None,
            "techpack_found": techpack_names,
            "warning":        "Tìm thấy techpack nhưng không trích xuất được trim items (có thể hết quota API)",
        }

    # Dedup cross-file
    before = len(all_trim_items)
    all_trim_items = extractor._deduplicate(all_trim_items)
    if len(all_trim_items) < before:
        logger.info(f"Cross-file dedup: {before} → {len(all_trim_items)}")

    # Export trimlist Excel
    trim_dir  = os.path.join(OUTPUT_DIR, "trimlist")
    os.makedirs(trim_dir, exist_ok=True)
    trim_path = os.path.join(trim_dir, f"trimlist_{timestamp}.xlsx")
    TrimlistExporter.export(trim_items=all_trim_items, output_path=trim_path, meta=meta)

    return {
        "status":           "success",
        "timestamp":        timestamp,
        "po":               po_data,
        "techpack_found":   techpack_names,
        "trimlist": {
            "item_count": len(all_trim_items),
            "trim_items": all_trim_items,
            "excel_path": trim_path,
        },
    }


@router.get("/download/po/{timestamp}")
def download_po(timestamp: str):
    path = os.path.join(OUTPUT_DIR, f"output_po_{timestamp}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(404, "File không tồn tại")
    return FileResponse(path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"PO_{timestamp}.xlsx")


@router.get("/download/trimlist/{timestamp}")
def download_trimlist(timestamp: str):
    path = os.path.join(OUTPUT_DIR, f"trimlist/trimlist_{timestamp}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(404, "File không tồn tại")
    return FileResponse(path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"Trimlist_{timestamp}.xlsx")


# ── Scan endpoints ─────────────────────────────────────────────────────────────

@router.get("/scan/status")
def scan_status():
    """
    Trả về trạng thái folder: tổng file, đã xử lý, file mới chưa xử lý.
    Frontend gọi endpoint này khi mở trang để biết có file mới không.
    """
    all_files  = get_all_po_files()
    new_files  = get_new_files()
    log        = get_processed_log()

    return {
        "scan_folder":   SCAN_FOLDER,
        "total_files":   len(all_files),
        "processed":     len(log),
        "new_files":     len(new_files),
        "new_filenames": [os.path.basename(f) for f in new_files],
        "log":           log,
    }


@router.get("/scan/test")
def scan_test():
    """Debug endpoint — kiểm tra folder và log hoạt động không."""
    try:
        all_files = get_all_po_files()
        new_files = get_new_files()
        log       = get_processed_log()
        return {
            "ok":         True,
            "scan_folder_exists": os.path.exists(SCAN_FOLDER),
            "total_files":  len(all_files),
            "new_files":    len(new_files),
            "log_entries":  len(log),
            "filenames":    [os.path.basename(f) for f in all_files],
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "trace": traceback.format_exc()}


@router.post("/scan/run")
async def scan_and_process():
    """
    Quét folder, tìm file mới, tự động xử lý từng file.
    Trả về kết quả tất cả file đã xử lý trong lần quét này.
    """
    try:
        new_files = get_new_files()
    except Exception as e:
        raise HTTPException(500, f"Lỗi đọc folder: {e}")

    if not new_files:
        return {
            "scanned": 0,
            "results": [],
            "message": "Không có file mới — tất cả đã được xử lý rồi.",
        }

    results = []
    for file_path in new_files:
        filename = os.path.basename(file_path)
        try:
            # Chạy blocking POAgent trong thread pool để không block event loop
            result = await asyncio.to_thread(_process_file_path_sync, file_path)
            mark_processed(file_path, result)
            results.append({"filename": filename, **result})
        except Exception as e:
            err = str(e)
            logger.error(f"Scan lỗi file {filename}: {err}\n{traceback.format_exc()}")
            mark_failed(file_path, err)
            results.append({
                "filename": filename,
                "status":   "error",
                "error":    err,
            })

    return {
        "scanned": len(new_files),
        "results": results,
        "message": f"Đã xử lý {len(new_files)} file mới.",
    }


@router.post("/scan/run-one")
async def scan_run_one(body: dict):
    """
    Xử lý MỘT file PO theo tên file. Frontend gọi lần lượt từng file.
    Body: {"filename": "test_po.pdf"}
    """
    filename = body.get("filename", "")
    if not filename:
        raise HTTPException(400, "Thiếu filename")

    # Tìm đường dẫn đầy đủ
    all_files = get_all_po_files()
    file_path = next((f for f in all_files if os.path.basename(f) == filename), None)
    if not file_path:
        raise HTTPException(404, f"Không tìm thấy file '{filename}' trong folder")

    try:
        result = await asyncio.to_thread(_process_file_path_sync, file_path)
        mark_processed(file_path, result)
        return {"filename": filename, **result}
    except Exception as e:
        err = str(e)
        logger.error(f"run-one lỗi {filename}: {err}\n{traceback.format_exc()}")
        mark_failed(file_path, err)
        return {"filename": filename, "status": "error", "error": err}


@router.post("/scan/reset/{filename}")
def scan_reset(filename: str):
    """Xóa file khỏi log để xử lý lại lần sau."""
    ok = reset_file(filename)
    if not ok:
        raise HTTPException(404, f"'{filename}' không có trong log")
    return {"message": f"Đã reset '{filename}' — sẽ được xử lý lại khi quét."}


@router.get("/trimlist-preview/{timestamp}")
def trimlist_preview(timestamp: str):
    """Đọc file Trimlist Excel và trả về danh sách trim items để preview."""
    import openpyxl
    path = os.path.join("sample_data", "trimlist", f"trimlist_{timestamp}.xlsx")
    if not os.path.exists(path):
        raise HTTPException(404, "File Trimlist không tồn tại")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # Tìm header row (chứa "Trim Item" hoặc "NO")
        header_idx = next(
            (i for i, r in enumerate(rows) if r and any(
                str(c or "").strip().upper() in ("NO", "TRIM ITEM", "TRIM_ITEM")
                for c in r
            )), None
        )
        if header_idx is None:
            return {"items": [], "warning": "Không tìm được header"}
        headers = [str(c or "").strip().lower().replace(" ", "_") for c in rows[header_idx]]
        items = []
        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            obj = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            # Bỏ dòng total
            name = str(obj.get("trim_item") or obj.get("trim item") or "").strip()
            if not name or name.upper() in ("TOTAL", "TỔNG CỘNG", "GRAND TOTAL"):
                continue
            items.append({
                "trim_item": name,
                "spec":      str(obj.get("spec") or obj.get("specification") or ""),
                "supplier":  str(obj.get("supplier") or obj.get("nha_cung_cap") or ""),
                "unit":      str(obj.get("unit") or obj.get("dvt") or ""),
                "total_qty": obj.get("total_qty") or obj.get("total qty") or obj.get("qty") or 0,
            })
        wb.close()
        return {"items": items[:20]}  # giới hạn 20 dòng preview
    except Exception as e:
        raise HTTPException(500, f"Lỗi đọc file: {e}")


class SendTrimlistEmailRequest(BaseModel):
    to_email:   str
    timestamp:  str
    po_number:  str = ""
    style_code: str = ""
    total_qty:  int = 0
    trim_count: int = 0


@router.post("/send-trimlist-email")
def send_trimlist_email_endpoint(body: SendTrimlistEmailRequest):
    """Gửi file Trimlist Excel qua Gmail."""
    from backend.utils.email_sender import send_trimlist_email

    trimlist_path = os.path.join("sample_data", "trimlist", f"trimlist_{body.timestamp}.xlsx")
    if not os.path.exists(trimlist_path):
        raise HTTPException(404, "File Trimlist không tồn tại")

    subject = f"[MCNA] Trim List — PO: {body.po_number or body.timestamp}"
    html_body = f"""
<div style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">
  <p>Kính gửi,</p>
  <p>Vui lòng xem file <strong>Trim List</strong> đính kèm.</p>
  <table style="border-collapse: collapse; margin: 16px 0;">
    <tr><td style="padding: 4px 12px 4px 0; color: #666;">PO Number:</td>
        <td><strong>{body.po_number or "—"}</strong></td></tr>
    <tr><td style="padding: 4px 12px 4px 0; color: #666;">Style Code:</td>
        <td><strong>{body.style_code or "—"}</strong></td></tr>
    <tr><td style="padding: 4px 12px 4px 0; color: #666;">Order Qty:</td>
        <td><strong>{body.total_qty:,} pcs</strong></td></tr>
    <tr><td style="padding: 4px 12px 4px 0; color: #666;">Số loại trim:</td>
        <td><strong>{body.trim_count}</strong></td></tr>
  </table>
  <p style="color: #999; font-size: 12px;">Email được gửi tự động từ hệ thống AI Agent — MCNA Garment</p>
</div>
"""
    try:
        send_trimlist_email(
            to_email=body.to_email,
            subject=subject,
            body=html_body,
            attachment_path=trimlist_path,
        )
    except Exception as e:
        raise HTTPException(500, f"Gửi email thất bại: {e}")

    return {"status": "success", "message": f"Đã gửi Trim List đến {body.to_email}"}


class SendTelegramRequest(BaseModel):
    chat_id:    str = ""
    timestamp:  str
    po_number:  str = ""
    style_code: str = ""
    total_qty:  int = 0
    trim_count: int = 0


@router.post("/send-trimlist-telegram")
def send_trimlist_telegram_endpoint(body: SendTelegramRequest):
    """Gửi file Trimlist Excel qua Telegram."""
    from backend.utils.telegram_sender import send_telegram_file
    from backend.config.settings import settings

    trimlist_path = os.path.join("sample_data", "trimlist", f"trimlist_{body.timestamp}.xlsx")
    if not os.path.exists(trimlist_path):
        raise HTTPException(404, "File Trimlist không tồn tại")

    chat_id = body.chat_id or settings.TELEGRAM_DEFAULT_CHAT_ID
    if not chat_id:
        raise HTTPException(400, "Chưa có chat_id")

    caption = (
        f"📋 <b>Trim List</b>\n"
        f"PO: <b>{body.po_number or '—'}</b>\n"
        f"Style: <b>{body.style_code or '—'}</b>\n"
        f"Qty: <b>{body.total_qty:,} pcs</b>\n"
        f"Số loại trim: <b>{body.trim_count}</b>\n"
        f"<i>MCNA Garment — AI Agent</i>"
    )
    try:
        send_telegram_file(chat_id=chat_id, file_path=trimlist_path, caption=caption)
    except Exception as e:
        raise HTTPException(500, f"Gửi Telegram thất bại: {e}")

    return {"status": "success", "message": f"Đã gửi Trim List qua Telegram"}


@router.get("/trimlist-pdf/{timestamp}")
def trimlist_pdf(timestamp: str):
    """Export file Trimlist thành PDF và trả về để download."""
    import openpyxl
    from fpdf import FPDF

    xlsx_path = os.path.join("sample_data", "trimlist", f"trimlist_{timestamp}.xlsx")
    if not os.path.exists(xlsx_path):
        raise HTTPException(404, "File Trimlist không tồn tại")

    # Đọc Excel
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Tìm header row
    header_idx = next(
        (i for i, r in enumerate(rows) if r and any(
            str(c or "").strip().upper() in ("NO", "TRIM ITEM", "TRIM_ITEM") for c in r
        )), None
    )

    # Lấy metadata (các dòng trước header)
    meta_lines = []
    if header_idx:
        for r in rows[:header_idx]:
            line = " | ".join(str(c or "").strip() for c in r if c)
            if line:
                meta_lines.append(line)

    # Lấy data rows
    headers = []
    data_rows = []
    if header_idx is not None:
        headers = [str(c or "").strip() for c in rows[header_idx]]
        for row in rows[header_idx + 1:]:
            if not any(row):
                continue
            cells = [str(c) if c is not None else "" for c in row]
            # Bỏ dòng total
            if cells and cells[0].strip().upper() in ("TOTAL", "TỔNG CỘNG", "GRAND TOTAL"):
                continue
            data_rows.append(cells)

    # Tạo PDF — dùng font Arial TTF để hỗ trợ tiếng Việt
    ARIAL    = r"C:\Windows\Fonts\arial.ttf"
    ARIAL_B  = r"C:\Windows\Fonts\arialbd.ttf"
    ARIAL_I  = r"C:\Windows\Fonts\ariali.ttf"

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("Arial",  "",  ARIAL)
    pdf.add_font("Arial",  "B", ARIAL_B)
    pdf.add_font("Arial",  "I", ARIAL_I)
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_margins(10, 10, 10)

    # Tiêu đề
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, "TRIM LIST", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Arial", "", 8)
    for line in meta_lines[:5]:
        pdf.cell(0, 5, line[:120], align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    if headers and data_rows:
        # Tính độ rộng cột theo số ký tự
        page_w = pdf.w - 20
        col_widths = [max(len(h), max((len(str(r[i])) if i < len(r) else 0 for r in data_rows), default=0))
                      for i, h in enumerate(headers)]
        total_chars = sum(col_widths) or 1
        col_w = [max(12, (page_w * w / total_chars)) for w in col_widths]
        actual_total = sum(col_w)
        if actual_total > page_w:
            col_w = [w * page_w / actual_total for w in col_w]

        # Header row
        pdf.set_font("Arial", "B", 7)
        pdf.set_fill_color(63, 84, 186)
        pdf.set_text_color(255, 255, 255)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 6, h[:30], border=1, fill=True)
        pdf.ln()

        # Data rows
        pdf.set_font("Arial", "", 7)
        pdf.set_text_color(0, 0, 0)
        for ri, row in enumerate(data_rows):
            fill = (ri % 2 == 0)
            pdf.set_fill_color(245, 246, 252) if fill else pdf.set_fill_color(255, 255, 255)
            for i in range(len(headers)):
                val = row[i] if i < len(row) else ""
                pdf.cell(col_w[i], 5.5, str(val)[:50], border=1, fill=fill)
            pdf.ln()
    else:
        pdf.set_font("Arial", "I", 9)
        pdf.cell(0, 8, "Không có dữ liệu", align="C", new_x="LMARGIN", new_y="NEXT")

    # Footer
    pdf.set_font("Arial", "I", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"MCNA Garment — AI Agent | Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
             align="R", new_x="LMARGIN", new_y="NEXT")

    # Lưu PDF
    pdf_dir = os.path.join("sample_data", "trimlist")
    pdf_path = os.path.join(pdf_dir, f"trimlist_{timestamp}.pdf")
    pdf.output(pdf_path)

    return FileResponse(
        pdf_path,
        media_type="application/pdf",
        filename=f"TrimList_{timestamp}.pdf",
    )


def _process_file_path_sync(file_path: str) -> dict:
    """Xử lý 1 file PO theo đường dẫn tuyệt đối (dùng nội bộ)."""
    ext = os.path.splitext(file_path)[-1].lower()
    ALLOWED = {".pdf", ".xlsx", ".xls", ".xlsm", ".docx", ".doc",
               ".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp",
               ".eml", ".msg"}
    if ext not in ALLOWED:
        raise ValueError(f"Định dạng '{ext}' không hỗ trợ")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # PO Agent
    agent  = POAgent()
    result = agent.process_request(
        user_request="Extract PO từ file và lưu vào database",
        file_path=file_path,
    )
    if result["status"] != "success":
        raise ValueError(f"Trích xuất PO thất bại: {result.get('reason', 'Unknown')}")

    results = result.get("results", {})
    header  = results.get("header_extractor", {}).get("header", {})
    items   = results.get("item_extractor",  {}).get("items",  [])

    combined   = {**header, "items": items,
                  "total_quantity_all": sum(i.get("total_quantity") or 0 for i in items),
                  "total_amount":       sum(i.get("total_price")    or 0.0 for i in items)}
    normalized = DataMapper.map_po_data(combined)
    po_model   = CanonicalSchema.validate_and_load(normalized)

    json_path  = os.path.join(OUTPUT_DIR, f"output_po_{timestamp}.json")
    excel_path = os.path.join(OUTPUT_DIR, f"output_po_{timestamp}.xlsx")
    JsonExporter.export(po_model, json_path)
    ExcelExporter.export(po_model, excel_path)

    po_id = None
    try:
        db    = SupabaseClient()
        po_id = db.insert_po(header=header, items=items)
    except Exception as e:
        logger.warning(f"Supabase bỏ qua: {e}")

    style_code = po_model.items[0].style_code if po_model.items else ""
    order_qty  = po_model.total_quantity_all or sum(i.total_quantity or 0 for i in po_model.items)
    _item0     = po_model.items[0] if po_model.items else None
    style_name = getattr(po_model, "style_name", "") or (getattr(_item0, "style_name", "") if _item0 else "")
    factory    = getattr(po_model, "factory", "") or ""

    po_data = {
        "po_id":        po_id,
        "po_number":    po_model.po_number,
        "style_code":   style_code,
        "style_name":   style_name,
        "buyer":        po_model.buyer,
        "factory":      factory,
        "total_qty":    order_qty,
        "total_amount": po_model.total_amount,
        "item_count":   len(po_model.items),
        "items":        [i.model_dump() for i in po_model.items],
        "excel_path":   excel_path,
    }

    # Tìm techpack & tạo trimlist
    techpack_files = _find_techpack(style_code)
    if not techpack_files:
        return {
            "status":         "partial",
            "timestamp":      timestamp,
            "po":             po_data,
            "trimlist":       None,
            "techpack_found": [],
            "warning":        f"Không tìm thấy techpack trong '{TECHPACK_DIR}/'",
        }

    meta = {
        "po_number":   po_model.po_number  or "",
        "style_code":  style_code,
        "style_name":  style_name,
        "buyer":       po_model.buyer      or "",
        "order_qty":   f"{order_qty:,} pcs",
        "factory":     factory,
        "season":      getattr(po_model, "season", "") or "",
        "date":        datetime.now().strftime("%d/%m/%Y"),
        "prepared_by": "AI Agent (Auto Scan)",
    }

    extractor      = TrimlistExtractor()
    all_trim_items = []
    techpack_names = []

    for tp_path in techpack_files:
        techpack_names.append(os.path.basename(tp_path))
        r = read_file(tp_path)
        if not r.get("success"):
            continue
        try:
            items_trim = extractor.extract(r["text"], order_qty=order_qty)
            for it in items_trim:
                it["_source_file"] = os.path.basename(tp_path)
            all_trim_items.extend(items_trim)
        except RuntimeError as e:
            if "RATE LIMIT" in str(e):
                raise
            logger.error(f"Trim extract lỗi: {e}")

    if not all_trim_items:
        return {
            "status":         "partial",
            "timestamp":      timestamp,
            "po":             po_data,
            "trimlist":       None,
            "techpack_found": techpack_names,
            "warning":        "Tìm thấy techpack nhưng không trích xuất được trim items",
        }

    before = len(all_trim_items)
    all_trim_items = extractor._deduplicate(all_trim_items)
    if len(all_trim_items) < before:
        logger.info(f"Dedup: {before} → {len(all_trim_items)}")

    trim_dir  = os.path.join(OUTPUT_DIR, "trimlist")
    os.makedirs(trim_dir, exist_ok=True)
    trim_path = os.path.join(trim_dir, f"trimlist_{timestamp}.xlsx")
    TrimlistExporter.export(trim_items=all_trim_items, output_path=trim_path, meta=meta)

    # Lưu trimlist vào Supabase
    trimlist_session_id = None
    try:
        from backend.database.supabase_client import SupabaseClient
        db = SupabaseClient()
        if not db.mock_mode:
            trimlist_session_id = db.insert_trimlist_session(
                po_id=po_data.get("po_id"),
                meta={**meta, "techpack_file": techpack_names[0] if techpack_names else ""},
                trim_items=all_trim_items,
                excel_path=trim_path,
            )
            logger.info(f"Saved trimlist_session id={trimlist_session_id}")
    except Exception as e:
        logger.warning(f"Supabase trimlist save bo qua: {e}")

    return {
        "status":         "success",
        "timestamp":      timestamp,
        "po":             po_data,
        "techpack_found": techpack_names,
        "trimlist": {
            "item_count":  len(all_trim_items),
            "trim_items":  all_trim_items,
            "excel_path":  trim_path,
            "session_id":  trimlist_session_id,
        },
    }
